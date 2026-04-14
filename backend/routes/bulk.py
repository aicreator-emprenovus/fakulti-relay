from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query
from fastapi.responses import StreamingResponse
import io
import logging
from datetime import datetime, timezone
from typing import Optional
from database import db
from auth import get_current_user
from utils import normalize_phone_ec, find_lead_by_phone, STAGE_LABELS
import uuid

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api")


@router.post("/bulk/upload")
async def bulk_upload(file: UploadFile = File(...), user=Depends(get_current_user)):
    import openpyxl

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [cell.value for cell in ws[1] if cell.value]
    header_map = {}
    for i, h in enumerate(headers):
        h_lower = h.lower().strip()
        if "nombre" in h_lower or "name" in h_lower:
            header_map["name"] = i
        elif "whatsapp" in h_lower or "telefono" in h_lower or "phone" in h_lower:
            header_map["whatsapp"] = i
        elif "ciudad" in h_lower or "city" in h_lower:
            header_map["city"] = i
        elif "producto" in h_lower or "product" in h_lower:
            header_map["product_interest"] = i
        elif "fecha" in h_lower or "date" in h_lower:
            header_map["purchase_date"] = i
        elif "email" in h_lower:
            header_map["email"] = i
        elif "fuente" in h_lower or "source" in h_lower:
            header_map["source"] = i
        elif "temporada" in h_lower or "season" in h_lower:
            header_map["season"] = i
        elif "canal" in h_lower or "channel" in h_lower:
            header_map["channel"] = i

    created = 0
    updated = 0
    errors = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            name = str(row[header_map.get("name", 0)] or "").strip()
            whatsapp_raw = str(row[header_map.get("whatsapp", 1)] or "").strip()
            if not name or not whatsapp_raw:
                continue
            whatsapp = normalize_phone_ec(whatsapp_raw)

            city = str(row[header_map.get("city", 2)] or "").strip() if "city" in header_map else ""
            product = str(row[header_map.get("product_interest", 3)] or "").strip() if "product_interest" in header_map else ""
            email = str(row[header_map.get("email", -1)] or "").strip() if "email" in header_map else ""
            source_val = str(row[header_map.get("source", -1)] or "").strip() if "source" in header_map else "Carga masiva"
            season_val = str(row[header_map.get("season", -1)] or "").strip() if "season" in header_map else ""
            channel_val = str(row[header_map.get("channel", -1)] or "").strip() if "channel" in header_map else ""

            has_purchase = "purchase_date" in header_map and row[header_map["purchase_date"]]
            stage = "cliente_nuevo" if has_purchase else "nuevo"

            existing = await find_lead_by_phone(whatsapp)
            if existing:
                update_data = {"name": name, "whatsapp": whatsapp, "last_interaction": datetime.now(timezone.utc).isoformat()}
                if city:
                    update_data["city"] = city
                if product:
                    update_data["product_interest"] = product
                if has_purchase:
                    update_data["funnel_stage"] = stage
                await db.leads.update_one({"id": existing["id"]}, {"$set": update_data})
                updated += 1
            else:
                lead_doc = {
                    "id": str(uuid.uuid4()),
                    "name": name, "whatsapp": whatsapp, "city": city, "email": email,
                    "product_interest": product, "source": source_val,
                    "season": season_val, "channel": channel_val,
                    "game_used": None, "prize_obtained": None,
                    "funnel_stage": stage, "status": "activo",
                    "purchase_history": [], "coupon_used": None, "recompra_date": None, "notes": "",
                    "last_interaction": datetime.now(timezone.utc).isoformat(),
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.leads.insert_one(lead_doc)
                created += 1
        except Exception as e:
            logger.error(f"Bulk upload row error: {e}")
            errors += 1

    return {"created": created, "updated": updated, "errors": errors, "total_processed": created + updated + errors}


@router.get("/bulk/download")
async def bulk_download(
    download_type: str = Query("all"),
    stage: Optional[str] = None,
    product: Optional[str] = None,
    user=Depends(get_current_user)
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    query = {}
    if download_type == "stage" and stage:
        query["funnel_stage"] = stage
    elif download_type == "product" and product:
        query["product_interest"] = {"$regex": product, "$options": "i"}
    elif download_type == "fibeca":
        query["source"] = "Fibeca"
    elif download_type == "game":
        query["game_used"] = {"$ne": None}
    elif download_type == "recompra":
        query["funnel_stage"] = {"$in": ["cliente_nuevo", "cliente_activo"]}

    leads = await db.leads.find(query, {"_id": 0}).to_list(10000)
    all_leads = await db.leads.find({}, {"_id": 0}).to_list(10000)
    products_list = await db.products.find({}, {"_id": 0}).to_list(100)

    wb = openpyxl.Workbook()

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A6B3C", end_color="1A6B3C", fill_type="solid")
    header_fill2 = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_fill3 = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB")
    )
    center_align = Alignment(horizontal="center", vertical="center")

    def style_header(ws, row, fill):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = fill
            cell.alignment = center_align
            cell.border = thin_border

    def style_data(ws, start_row):
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 35)

    stage_labels = STAGE_LABELS

    # Sheet 1: Base de Datos
    ws1 = wb.active
    ws1.title = "Base de Datos"
    headers_row = ["#", "Nombre", "WhatsApp", "Email", "Ciudad", "Producto de Interes", "Fuente", "Etapa del Embudo", "Estado", "Juego Usado", "Premio Obtenido", "Cupon", "Ultima Interaccion", "Fecha de Registro"]
    ws1.append(headers_row)
    style_header(ws1, 1, header_fill)
    for i, lead in enumerate(leads, 1):
        ws1.append([
            i, lead.get("name", ""), lead.get("whatsapp", ""), lead.get("email", ""),
            lead.get("city", ""), lead.get("product_interest", ""), lead.get("source", ""),
            stage_labels.get(lead.get("funnel_stage", ""), lead.get("funnel_stage", "")),
            lead.get("status", ""), lead.get("game_used", "") or "", lead.get("prize_obtained", "") or "",
            lead.get("coupon_used", "") or "", (lead.get("last_interaction", "") or "")[:19],
            (lead.get("created_at", "") or "")[:19]
        ])
    style_data(ws1, 2)
    auto_width(ws1)
    ws1.freeze_panes = "B2"
    ws1.auto_filter.ref = f"A1:N{len(leads)+1}"

    # Sheet 2: Embudo
    ws2 = wb.create_sheet("Embudo de Ventas")
    ws2.append(["Etapa del Embudo", "Cantidad de Leads", "Porcentaje", "Tasa de Conversion"])
    style_header(ws2, 1, header_fill2)
    stage_order = ["nuevo", "interesado", "en_negociacion", "cliente_nuevo", "cliente_activo", "perdido"]
    total_leads = len(all_leads)
    for s in stage_order:
        count = sum(1 for lead in all_leads if lead.get("funnel_stage") == s)
        pct = f"{(count/total_leads*100):.1f}%" if total_leads > 0 else "0%"
        conv = ""
        if s in ("cliente_nuevo", "cliente_activo"):
            conv = f"{(count/total_leads*100):.1f}%" if total_leads > 0 else "0%"
        ws2.append([stage_labels.get(s, s), count, pct, conv])
    ws2.append([])
    ws2.append(["TOTAL LEADS", total_leads, "100%", ""])
    ws2[ws2.max_row][0].font = Font(bold=True)
    ws2[ws2.max_row][1].font = Font(bold=True)
    style_data(ws2, 2)
    auto_width(ws2)

    # Sheet 3: Fuentes
    ws3 = wb.create_sheet("Fuentes de Trafico")
    ws3.append(["Fuente", "Total Leads", "Porcentaje", "Clientes Convertidos", "Tasa de Conversion"])
    style_header(ws3, 1, header_fill3)
    sources = {}
    for lead in all_leads:
        src = lead.get("source", "Desconocido") or "Desconocido"
        if src not in sources:
            sources[src] = {"total": 0, "clientes": 0}
        sources[src]["total"] += 1
        if lead.get("funnel_stage") in ["cliente_nuevo", "cliente_activo"]:
            sources[src]["clientes"] += 1
    for src, data in sorted(sources.items(), key=lambda x: x[1]["total"], reverse=True):
        pct = f"{(data['total']/total_leads*100):.1f}%" if total_leads > 0 else "0%"
        conv = f"{(data['clientes']/data['total']*100):.1f}%" if data["total"] > 0 else "0%"
        ws3.append([src, data["total"], pct, data["clientes"], conv])
    style_data(ws3, 2)
    auto_width(ws3)

    # Sheet 4: Productos
    ws4 = wb.create_sheet("Catalogo de Productos")
    ws4.append(["Producto", "Codigo", "Precio", "Precio Original", "Stock", "Categoria", "Leads Interesados"])
    style_header(ws4, 1, header_fill)
    for p in products_list:
        interested = sum(1 for lead in all_leads if p["name"].lower() in (lead.get("product_interest", "") or "").lower())
        ws4.append([p["name"], p.get("code", ""), f"${p['price']}", f"${p.get('original_price', '')}", p.get("stock", ""), p.get("category", ""), interested])
    style_data(ws4, 2)
    auto_width(ws4)

    # Sheet 5: Ciudades
    ws5 = wb.create_sheet("Leads por Ciudad")
    ws5.append(["Ciudad", "Total Leads", "Porcentaje", "Clientes"])
    style_header(ws5, 1, header_fill2)
    cities = {}
    for lead in all_leads:
        city = lead.get("city", "Sin ciudad") or "Sin ciudad"
        if city not in cities:
            cities[city] = {"total": 0, "clientes": 0}
        cities[city]["total"] += 1
        if lead.get("funnel_stage") in ["cliente_nuevo", "cliente_activo"]:
            cities[city]["clientes"] += 1
    for city, data in sorted(cities.items(), key=lambda x: x[1]["total"], reverse=True):
        pct = f"{(data['total']/total_leads*100):.1f}%" if total_leads > 0 else "0%"
        ws5.append([city, data["total"], pct, data["clientes"]])
    style_data(ws5, 2)
    auto_width(ws5)

    # Sheet 6: Periodos
    ws6 = wb.create_sheet("Leads por Periodo")
    ws6.append(["Periodo", "Nuevos Leads", "Leads Interesados", "En Negociacion", "Clientes Nuevos", "Perdidos", "Total"])
    style_header(ws6, 1, header_fill3)
    months = {}
    for lead in all_leads:
        created = lead.get("created_at", "")
        if created:
            month_key = created[:7]
            if month_key not in months:
                months[month_key] = {"nuevo": 0, "interesado": 0, "en_negociacion": 0, "cliente_nuevo": 0, "perdido": 0, "total": 0}
            stage_val = lead.get("funnel_stage", "nuevo")
            if stage_val in months[month_key]:
                months[month_key][stage_val] += 1
            months[month_key]["total"] += 1
    for period in sorted(months.keys()):
        d = months[period]
        ws6.append([period, d["nuevo"], d["interesado"], d["en_negociacion"], d["cliente_nuevo"], d["perdido"], d["total"]])
    style_data(ws6, 2)
    auto_width(ws6)

    # Sheet 7: Resumen Ejecutivo
    ws7 = wb.create_sheet("Resumen Ejecutivo")
    ws7.sheet_properties.tabColor = "FFD700"
    title_font = Font(name="Calibri", bold=True, size=16, color="1A6B3C")
    subtitle_font = Font(name="Calibri", bold=True, size=12, color="333333")
    ws7.append(["REPORTE EJECUTIVO - FAKULTI LABORATORIOS"])
    ws7["A1"].font = title_font
    ws7.merge_cells("A1:D1")
    ws7.append([f"Generado: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} UTC"])
    ws7["A2"].font = Font(italic=True, color="666666")
    ws7.append([])
    ws7.append(["KPI", "Valor"])
    ws7["A4"].font = subtitle_font
    ws7["B4"].font = subtitle_font
    clientes = sum(1 for lead in all_leads if lead.get("funnel_stage") in ["cliente_nuevo", "cliente_activo"])
    interesados = sum(1 for lead in all_leads if lead.get("funnel_stage") == "interesado")
    negociacion = sum(1 for lead in all_leads if lead.get("funnel_stage") == "en_negociacion")
    perdidos = sum(1 for lead in all_leads if lead.get("funnel_stage") == "perdido")
    kpis = [
        ("Total Leads", total_leads), ("Clientes Activos", clientes),
        ("Leads Interesados", interesados), ("En Negociacion", negociacion),
        ("Leads Perdidos", perdidos),
        ("Tasa de Conversion", f"{(clientes/total_leads*100):.1f}%" if total_leads > 0 else "0%"),
        ("Fuente Principal", max(sources.items(), key=lambda x: x[1]["total"])[0] if sources else "N/A"),
        ("Ciudad Principal", max(cities.items(), key=lambda x: x[1]["total"])[0] if cities else "N/A"),
    ]
    for kpi, val in kpis:
        ws7.append([kpi, val])
    style_data(ws7, 4)
    auto_width(ws7)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"fakulti_reporte_{download_type}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})
