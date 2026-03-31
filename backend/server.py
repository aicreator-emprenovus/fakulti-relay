from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Query, Header
from fastapi.responses import StreamingResponse, FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
from passlib.context import CryptContext
import io
import json
import secrets
import re

import httpx

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

JWT_SECRET = os.environ.get('JWT_SECRET', 'faculty-crm-jwt-secret-2024')
JWT_ALGORITHM = "HS256"
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

def safe_hash_password(password: str) -> str:
    """Hash password with fallback to direct bcrypt if passlib fails."""
    try:
        return pwd_context.hash(password)
    except Exception:
        import bcrypt as _bcrypt
        return _bcrypt.hashpw(password.encode("utf-8"), _bcrypt.gensalt()).decode("utf-8")

def safe_verify_password(plain: str, hashed: str) -> bool:
    """Verify password with fallback to direct bcrypt if passlib fails."""
    try:
        return pwd_context.verify(plain, hashed)
    except Exception:
        import bcrypt as _bcrypt
        return _bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

FUNNEL_STAGES = ["nuevo", "interesado", "en_negociacion", "cliente_nuevo", "cliente_activo", "perdido"]
STAGE_LABELS = {
    "nuevo": "Contacto inicial",
    "interesado": "Chat",
    "en_negociacion": "En Negociación",
    "cliente_nuevo": "Leads ganados",
    "cliente_activo": "Cartera activa",
    "perdido": "Perdido",
}
SOURCES = ["TV", "QR", "Fibeca", "pauta_digital", "web", "referido", "otro"]
SEASONS = ["verano", "invierno", "todo_el_año"]

def normalize_phone_ec(phone: str) -> str:
    """Normalize Ecuador phone to local format without +593."""
    phone = re.sub(r'[\s\-\(\)]', '', phone.strip())
    if phone.startswith('+593'):
        phone = '0' + phone[4:]
    elif phone.startswith('593') and len(phone) > 9:
        phone = '0' + phone[3:]
    return phone

def phone_to_international(phone: str) -> str:
    """Convert local Ecuador phone to international format for WhatsApp API (593XXXXXXXXX)."""
    phone = re.sub(r'[\s\-\(\)]', '', phone.strip())
    if phone.startswith('+'):
        return phone[1:]
    if phone.startswith('0'):
        return '593' + phone[1:]
    if phone.startswith('593'):
        return phone
    return '593' + phone

def phone_variants(phone: str) -> list:
    """Return all possible format variants for a phone number for flexible lookup."""
    normalized = normalize_phone_ec(phone)
    international = phone_to_international(phone)
    variants = list(set([normalized, international, phone.strip()]))
    return variants

async def find_lead_by_phone(phone: str):
    """Find a lead by phone number, checking all format variants."""
    variants = phone_variants(phone)
    lead = await db.leads.find_one({"whatsapp": {"$in": variants}}, {"_id": 0})
    return lead

# ========== PYDANTIC MODELS ==========

class LoginRequest(BaseModel):
    email: str
    password: str

class RegisterRequest(BaseModel):
    email: str
    password: str
    name: str

class LeadCreate(BaseModel):
    name: str
    whatsapp: str
    city: Optional[str] = ""
    email: Optional[str] = ""
    product_interest: Optional[str] = ""
    source: Optional[str] = "web"
    notes: Optional[str] = ""
    funnel_stage: Optional[str] = "nuevo"
    season: Optional[str] = ""
    channel: Optional[str] = ""

class LeadUpdate(BaseModel):
    name: Optional[str] = None
    whatsapp: Optional[str] = None
    city: Optional[str] = None
    email: Optional[str] = None
    product_interest: Optional[str] = None
    source: Optional[str] = None
    notes: Optional[str] = None
    funnel_stage: Optional[str] = None
    status: Optional[str] = None
    recompra_date: Optional[str] = None
    season: Optional[str] = None
    channel: Optional[str] = None
    assigned_advisor: Optional[str] = None
    ci_ruc: Optional[str] = None
    address: Optional[str] = None

class AdvisorCreate(BaseModel):
    name: str
    email: str
    password: str
    whatsapp: Optional[str] = ""
    status: Optional[str] = "disponible"
    specialization: Optional[str] = ""

class ProductCreate(BaseModel):
    name: str
    code: Optional[str] = ""
    description: Optional[str] = ""
    price: float
    original_price: Optional[float] = None
    image_url: Optional[str] = ""
    stock: Optional[int] = 100
    category: Optional[str] = "general"
    active: Optional[bool] = True

class GameConfigCreate(BaseModel):
    game_type: str
    name: str
    prizes: List[dict]
    active: Optional[bool] = True
    max_plays_per_whatsapp: Optional[int] = 1

class GamePlayRequest(BaseModel):
    game_type: str
    whatsapp: str
    name: str
    city: Optional[str] = ""

class QuotationCreate(BaseModel):
    lead_id: str
    items: List[dict]
    notes: Optional[str] = ""

class LoyaltySequenceCreate(BaseModel):
    product_id: str
    product_name: str
    messages: List[dict]
    active: Optional[bool] = True

class ChatMessageRequest(BaseModel):
    lead_id: Optional[str] = None
    session_id: str
    message: str

class PurchaseAdd(BaseModel):
    product_id: str
    product_name: str
    quantity: int = 1
    price: float

class AutomationRuleCreate(BaseModel):
    name: str
    trigger_type: str
    trigger_value: Optional[str] = ""
    action_type: str
    action_value: Optional[str] = ""
    description: Optional[str] = ""
    active: Optional[bool] = True

class WhatsAppConfigUpdate(BaseModel):
    phone_number_id: Optional[str] = ""
    access_token: Optional[str] = ""
    verify_token: Optional[str] = "fakulti-whatsapp-verify-token"
    business_name: Optional[str] = "Fakulti Laboratorios"

class AIConfigUpdate(BaseModel):
    intent_analysis: Optional[bool] = True
    lead_classification: Optional[bool] = True
    product_recommendation: Optional[bool] = True
    suggested_responses: Optional[bool] = True

class CRMWhatsAppReply(BaseModel):
    lead_id: str
    message: str

class QRCampaignCreate(BaseModel):
    name: str
    channel: str
    source: str
    product: Optional[str] = ""
    initial_message: str
    description: Optional[str] = ""
    active: Optional[bool] = True

# ========== AUTH UTILITIES ==========

def create_token(user_id: str, email: str):
    payload = {"user_id": user_id, "email": email, "exp": datetime.now(timezone.utc) + timedelta(hours=24)}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(authorization: Optional[str] = Header(None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Token requerido")
    token = authorization.split(" ")[1]
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.admin_users.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Usuario no encontrado")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalido")

# ========== AUTH ROUTES ==========

@api_router.post("/auth/login")
async def login(req: LoginRequest):
    user = await db.admin_users.find_one({"email": req.email}, {"_id": 0})
    if not user or not safe_verify_password(req.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Credenciales incorrectas")
    token = create_token(user["id"], user["email"])
    return {"token": token, "user": {"id": user["id"], "email": user["email"], "name": user["name"], "role": user.get("role", "admin")}}

@api_router.post("/auth/register")
async def register(req: RegisterRequest):
    existing = await db.admin_users.find_one({"email": req.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email ya registrado")
    user_doc = {
        "id": str(uuid.uuid4()),
        "email": req.email,
        "password_hash": safe_hash_password(req.password),
        "name": req.name,
        "role": "admin",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.admin_users.insert_one(user_doc)
    token = create_token(user_doc["id"], user_doc["email"])
    return {"token": token, "user": {"id": user_doc["id"], "email": user_doc["email"], "name": user_doc["name"], "role": "admin"}}

@api_router.get("/auth/me")
async def get_me(user=Depends(get_current_user)):
    return {"id": user["id"], "email": user["email"], "name": user["name"], "role": user.get("role", "admin")}

# ========== PASSWORD RESET REQUESTS ==========

class PasswordResetRequest(BaseModel):
    email: str

@api_router.post("/auth/forgot-password")
async def forgot_password(req: PasswordResetRequest):
    """Creates a password reset request. Shows a message that the request was sent to the appropriate authority."""
    user = await db.admin_users.find_one({"email": req.email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Email no encontrado en el sistema")
    
    role = user.get("role", "admin")
    
    # Determine who receives the notification
    if role == "admin":
        notify_role = "developer"
        display_message = "La solicitud fue enviada al desarrollador del sistema"
    elif role == "advisor":
        notify_role = "admin"
        display_message = "La solicitud fue enviada al Administrador del sistema"
    else:
        raise HTTPException(status_code=400, detail="Los desarrolladores no pueden solicitar reset por esta vía")
    
    # Store the reset request
    await db.password_reset_requests.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "user_email": user["email"],
        "user_name": user.get("name", ""),
        "user_role": role,
        "notify_role": notify_role,
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": display_message}

@api_router.get("/auth/password-reset-requests")
async def get_password_reset_requests(user=Depends(get_current_user)):
    """Get pending password reset requests. Developer sees admin requests, Admin sees advisor requests."""
    role = user.get("role", "admin")
    if role == "developer":
        query = {"notify_role": "developer", "status": "pending"}
    elif role == "admin":
        query = {"notify_role": "admin", "status": "pending"}
    else:
        return []
    requests = await db.password_reset_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(50)
    return requests

class ResetPasswordAction(BaseModel):
    new_password: str

@api_router.post("/auth/approve-reset/{request_id}")
async def approve_password_reset(request_id: str, user=Depends(get_current_user)):
    """Approve a password reset request. Developer approves admin, Admin approves advisor. No password needed — the user sets it themselves."""
    role = user.get("role", "admin")
    reset_req = await db.password_reset_requests.find_one({"id": request_id, "status": "pending"}, {"_id": 0})
    if not reset_req:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada")
    if role == "developer" and reset_req["user_role"] != "admin":
        raise HTTPException(status_code=403, detail="Sin permisos")
    if role == "admin" and reset_req["user_role"] != "advisor":
        raise HTTPException(status_code=403, detail="Sin permisos")
    
    reset_token = secrets.token_urlsafe(32)
    await db.password_reset_requests.update_one(
        {"id": request_id},
        {"$set": {"status": "approved", "reset_token": reset_token, "approved_by": user["id"], "approved_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": f"Solicitud aprobada. {reset_req['user_name']} podrá crear su nueva contraseña desde el login."}

@api_router.post("/auth/check-reset")
async def check_reset_available(body: dict):
    """Check if an approved reset exists for this email (no auth required)."""
    email = body.get("email", "")
    if not email:
        raise HTTPException(status_code=400, detail="Email requerido")
    request = await db.password_reset_requests.find_one(
        {"user_email": email, "status": "approved"},
        {"_id": 0, "id": 1, "user_name": 1}
    )
    if request:
        return {"has_approved_reset": True, "user_name": request.get("user_name", "")}
    return {"has_approved_reset": False}

@api_router.post("/auth/set-new-password")
async def set_new_password(body: dict):
    """User sets their own new password after reset was approved (no auth required)."""
    email = body.get("email", "")
    new_password = body.get("new_password", "")
    if not email or not new_password:
        raise HTTPException(status_code=400, detail="Email y nueva contraseña requeridos")
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 6 caracteres")
    
    request = await db.password_reset_requests.find_one(
        {"user_email": email, "status": "approved"}, {"_id": 0}
    )
    if not request:
        raise HTTPException(status_code=404, detail="No hay solicitud de restablecimiento aprobada para este email")
    
    await db.admin_users.update_one(
        {"id": request["user_id"]},
        {"$set": {"password_hash": safe_hash_password(new_password)}}
    )
    await db.password_reset_requests.update_one(
        {"id": request["id"]},
        {"$set": {"status": "resolved", "resolved_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": "Contraseña actualizada exitosamente. Ya puedes iniciar sesión."}

@api_router.post("/auth/reset-password/{request_id}")
async def execute_password_reset(request_id: str, body: ResetPasswordAction, user=Depends(get_current_user)):
    """Legacy: Execute a password reset with direct password (kept for admin resetting advisors)."""
    role = user.get("role", "admin")
    reset_req = await db.password_reset_requests.find_one({"id": request_id, "status": "pending"}, {"_id": 0})
    if not reset_req:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada")
    if role == "admin" and reset_req["user_role"] != "advisor":
        raise HTTPException(status_code=403, detail="Sin permisos")
    
    await db.admin_users.update_one(
        {"id": reset_req["user_id"]},
        {"$set": {"password_hash": safe_hash_password(body.new_password)}}
    )
    await db.password_reset_requests.update_one(
        {"id": request_id},
        {"$set": {"status": "resolved", "resolved_by": user["id"], "resolved_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": f"Contraseña de {reset_req['user_email']} restablecida exitosamente"}

@api_router.post("/auth/reset-password-direct")
async def direct_password_reset(body: dict, user=Depends(get_current_user)):
    """Direct password reset (Admin resets advisor only)."""
    role = user.get("role", "admin")
    target_user_id = body.get("user_id", "")
    new_password = body.get("new_password", "")
    if not target_user_id or not new_password:
        raise HTTPException(status_code=400, detail="user_id y new_password requeridos")
    
    target = await db.admin_users.find_one({"id": target_user_id}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    target_role = target.get("role", "admin")
    if role != "admin" or target_role != "advisor":
        raise HTTPException(status_code=403, detail="Solo el admin puede resetear contraseñas de asesores")
    
    await db.admin_users.update_one(
        {"id": target_user_id},
        {"$set": {"password_hash": safe_hash_password(new_password)}}
    )
    return {"message": f"Contraseña de {target['email']} restablecida exitosamente"}

# ========== DASHBOARD ROUTES ==========

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(user=Depends(get_current_user)):
    total_leads = await db.leads.count_documents({})
    stages = {}
    for stage in FUNNEL_STAGES:
        stages[stage] = await db.leads.count_documents({"funnel_stage": stage})
    
    total_sales = 0
    pipeline = [{"$unwind": "$purchase_history"}, {"$group": {"_id": None, "total": {"$sum": "$purchase_history.price"}, "count": {"$sum": 1}}}]
    result = await db.leads.aggregate(pipeline).to_list(1)
    if result:
        total_sales = result[0]["total"]
        total_orders = result[0]["count"]
    else:
        total_orders = 0
    
    clients = await db.leads.count_documents({"funnel_stage": {"$in": ["cliente_nuevo", "cliente_activo"]}})
    
    game_plays = await db.game_plays.count_documents({})
    game_conversions = await db.game_plays.count_documents({"converted": True})
    
    product_stats = await db.leads.aggregate([
        {"$unwind": "$purchase_history"},
        {"$group": {"_id": "$purchase_history.product_name", "count": {"$sum": 1}, "revenue": {"$sum": "$purchase_history.price"}}}
    ]).to_list(100)
    
    source_stats = await db.leads.aggregate([
        {"$group": {"_id": "$source", "count": {"$sum": 1}}}
    ]).to_list(100)
    
    recent_leads = await db.leads.find({}, {"_id": 0, "id": 1, "name": 1, "whatsapp": 1, "funnel_stage": 1, "product_interest": 1, "created_at": 1, "last_interaction": 1}).sort("created_at", -1).limit(10).to_list(10)
    
    return {
        "total_leads": total_leads,
        "stages": stages,
        "total_sales": round(total_sales, 2),
        "total_orders": total_orders,
        "total_clients": clients,
        "game_plays": game_plays,
        "game_conversions": game_conversions,
        "conversion_rate": round((clients / total_leads * 100) if total_leads > 0 else 0, 1),
        "product_stats": [{"name": p["_id"], "count": p["count"], "revenue": round(p["revenue"], 2)} for p in product_stats],
        "source_stats": [{"name": s["_id"] or "Sin fuente", "count": s["count"]} for s in source_stats],
        "recent_leads": recent_leads
    }

# ========== LEAD ROUTES ==========

@api_router.get("/leads")
async def get_leads(
    stage: Optional[str] = None,
    search: Optional[str] = None,
    source: Optional[str] = None,
    status: Optional[str] = None,
    season: Optional[str] = None,
    channel: Optional[str] = None,
    advisor: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    user=Depends(get_current_user)
):
    query = {}
    # Role-based filtering: advisors only see their assigned leads
    user_role = user.get("role", "admin")
    if user_role == "advisor":
        query["assigned_advisor"] = user["id"]
    elif advisor:
        query["assigned_advisor"] = advisor
    if stage:
        query["funnel_stage"] = stage
    if source:
        query["source"] = source
    if status:
        query["status"] = status
    if season:
        query["season"] = season
    if channel:
        query["channel"] = channel
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"whatsapp": {"$regex": search, "$options": "i"}},
            {"city": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}}
        ]
    
    total = await db.leads.count_documents(query)
    skip = (page - 1) * limit
    leads = await db.leads.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"leads": leads, "total": total, "page": page, "pages": (total + limit - 1) // limit}

@api_router.get("/leads/{lead_id}")
async def get_lead(lead_id: str, user=Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    if lead.get("assigned_advisor"):
        advisor = await db.admin_users.find_one({"id": lead["assigned_advisor"], "role": "advisor"}, {"_id": 0, "name": 1})
        lead["_advisor_name"] = advisor["name"] if advisor else ""
    return lead

@api_router.post("/leads")
async def create_lead(req: LeadCreate, user=Depends(get_current_user)):
    lead_data = req.model_dump()
    lead_data["whatsapp"] = normalize_phone_ec(lead_data["whatsapp"])
    lead_doc = {
        "id": str(uuid.uuid4()),
        **lead_data,
        "status": "activo",
        "game_used": None,
        "prize_obtained": None,
        "purchase_history": [],
        "coupon_used": None,
        "recompra_date": None,
        "last_interaction": datetime.now(timezone.utc).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.leads.insert_one(lead_doc)
    lead_doc.pop("_id", None)
    return lead_doc

@api_router.put("/leads/{lead_id}")
async def update_lead(lead_id: str, req: LeadUpdate, user=Depends(get_current_user)):
    update_data = {k: v for k, v in req.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No hay datos para actualizar")
    update_data["last_interaction"] = datetime.now(timezone.utc).isoformat()
    await db.leads.update_one({"id": lead_id}, {"$set": update_data})
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    return lead

@api_router.delete("/leads/{lead_id}")
async def delete_lead(lead_id: str, user=Depends(get_current_user)):
    result = await db.leads.delete_one({"id": lead_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    return {"message": "Lead eliminado"}

@api_router.post("/leads/{lead_id}/purchase")
async def add_purchase(lead_id: str, req: PurchaseAdd, user=Depends(get_current_user)):
    purchase = {
        "id": str(uuid.uuid4()),
        "product_id": req.product_id,
        "product_name": req.product_name,
        "quantity": req.quantity,
        "price": req.price * req.quantity,
        "date": datetime.now(timezone.utc).isoformat()
    }
    await db.leads.update_one(
        {"id": lead_id},
        {
            "$push": {"purchase_history": purchase},
            "$set": {
                "funnel_stage": "cliente_nuevo",
                "last_interaction": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    # Auto-enroll in loyalty sequence if one exists for this product
    seq = await db.loyalty_sequences.find_one({"product_id": req.product_id, "active": True}, {"_id": 0})
    if seq:
        existing_enrollment = await db.loyalty_enrollments.find_one({"lead_id": lead_id, "sequence_id": seq["id"]})
        if not existing_enrollment:
            lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
            now = datetime.now(timezone.utc)
            messages_schedule = []
            for msg in seq.get("messages", []):
                if msg.get("active", True):
                    send_date = now + timedelta(days=msg.get("day", 1))
                    messages_schedule.append({
                        "day": msg["day"],
                        "content": msg["content"],
                        "scheduled_date": send_date.isoformat(),
                        "status": "pendiente"
                    })
            enrollment = {
                "id": str(uuid.uuid4()),
                "lead_id": lead_id,
                "lead_name": lead.get("name", "") if lead else "",
                "lead_whatsapp": lead.get("whatsapp", "") if lead else "",
                "sequence_id": seq["id"],
                "sequence_name": seq.get("product_name", ""),
                "messages": messages_schedule,
                "status": "activo",
                "enrolled_at": now.isoformat(),
                "completed_at": None
            }
            await db.loyalty_enrollments.insert_one(enrollment)
            logger.info(f"Auto-enrolled lead {lead_id} in loyalty sequence {seq['id']}")
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    return lead

@api_router.put("/leads/{lead_id}/stage")
async def update_lead_stage(lead_id: str, stage: str = Query(...)):
    if stage not in FUNNEL_STAGES:
        raise HTTPException(status_code=400, detail="Etapa invalida")
    await db.leads.update_one({"id": lead_id}, {"$set": {"funnel_stage": stage, "last_interaction": datetime.now(timezone.utc).isoformat()}})
    return {"message": "Etapa actualizada"}

# ========== ADVISOR ROUTES ==========

@api_router.get("/advisors")
async def get_advisors(user=Depends(get_current_user)):
    """Get all advisors. Admin only."""
    advisors = await db.admin_users.find({"role": "advisor"}, {"_id": 0, "password_hash": 0}).to_list(100)
    # Add assigned leads count
    for a in advisors:
        a["leads_count"] = await db.leads.count_documents({"assigned_advisor": a["id"]})
        a["active_chats"] = await db.leads.count_documents({"assigned_advisor": a["id"], "bot_paused": True})
    return advisors

@api_router.post("/advisors")
async def create_advisor(req: AdvisorCreate, user=Depends(get_current_user)):
    """Create a new advisor. Admin only."""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden crear asesores")
    existing = await db.admin_users.find_one({"email": req.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email ya registrado")
    advisor_doc = {
        "id": str(uuid.uuid4()),
        "email": req.email,
        "password_hash": safe_hash_password(req.password),
        "name": req.name,
        "whatsapp": normalize_phone_ec(req.whatsapp) if req.whatsapp else "",
        "role": "advisor",
        "status": req.status or "disponible",
        "specialization": req.specialization or "",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.admin_users.insert_one(advisor_doc)
    advisor_doc.pop("_id", None)
    advisor_doc.pop("password_hash", None)
    return advisor_doc

@api_router.put("/advisors/{advisor_id}")
async def update_advisor(advisor_id: str, user=Depends(get_current_user)):
    """Update advisor. Admin only."""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    # Read JSON body manually for flexibility
    from starlette.requests import Request
    return {"message": "Use the specific update endpoint"}

@api_router.put("/advisors/{advisor_id}/status")
async def update_advisor_status(advisor_id: str, body: dict, user=Depends(get_current_user)):
    """Update advisor status/info. Admin only."""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    update = {}
    if "status" in body:
        update["status"] = body["status"]
    if "name" in body:
        update["name"] = body["name"]
    if "whatsapp" in body:
        update["whatsapp"] = normalize_phone_ec(body["whatsapp"]) if body["whatsapp"] else ""
    if "specialization" in body:
        update["specialization"] = body["specialization"]
    if "password" in body and body["password"]:
        update["password_hash"] = safe_hash_password(body["password"])
    if not update:
        raise HTTPException(status_code=400, detail="Sin campos para actualizar")
    await db.admin_users.update_one({"id": advisor_id, "role": "advisor"}, {"$set": update})
    advisor = await db.admin_users.find_one({"id": advisor_id}, {"_id": 0, "password_hash": 0})
    return advisor

@api_router.delete("/advisors/{advisor_id}")
async def delete_advisor(advisor_id: str, user=Depends(get_current_user)):
    """Delete advisor. Admin only."""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    # Unassign all leads from this advisor
    await db.leads.update_many({"assigned_advisor": advisor_id}, {"$set": {"assigned_advisor": ""}})
    await db.admin_users.delete_one({"id": advisor_id, "role": "advisor"})
    return {"message": "Asesor eliminado y leads desasignados"}

@api_router.put("/leads/{lead_id}/assign")
async def assign_lead_to_advisor(lead_id: str, body: dict, user=Depends(get_current_user)):
    """Assign a lead to an advisor."""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden asignar leads")
    advisor_id = body.get("advisor_id", "")
    if advisor_id:
        advisor = await db.admin_users.find_one({"id": advisor_id, "role": "advisor"}, {"_id": 0, "name": 1})
        if not advisor:
            raise HTTPException(status_code=404, detail="Asesor no encontrado")
    await db.leads.update_one(
        {"id": lead_id},
        {"$set": {"assigned_advisor": advisor_id, "needs_advisor": False, "last_interaction": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": f"Lead asignado a {advisor_id or 'sin asesor'}"}

@api_router.get("/advisors/notifications")
async def get_advisor_notifications(user=Depends(get_current_user)):
    """Get unread notifications for the current advisor (or all for admin)."""
    query = {"read": False}
    if user.get("role") == "advisor":
        query["advisor_id"] = user["id"]
    notifications = await db.advisor_notifications.find(query, {"_id": 0}).sort("created_at", -1).to_list(50)
    return notifications

@api_router.put("/advisors/notifications/{notif_id}/read")
async def mark_notification_read(notif_id: str, user=Depends(get_current_user)):
    await db.advisor_notifications.update_one({"id": notif_id}, {"$set": {"read": True}})
    return {"message": "Notificación leída"}

@api_router.put("/advisors/notifications/read-all")
async def mark_all_notifications_read(user=Depends(get_current_user)):
    query = {}
    if user.get("role") == "advisor":
        query["advisor_id"] = user["id"]
    await db.advisor_notifications.update_many(query, {"$set": {"read": True}})
    return {"message": "Todas las notificaciones leídas"}

# ========== PRODUCT ROUTES ==========

@api_router.get("/products")
async def get_products():
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    return products

@api_router.post("/products")
async def create_product(req: ProductCreate, user=Depends(get_current_user)):
    doc = {"id": str(uuid.uuid4()), **req.model_dump(), "created_at": datetime.now(timezone.utc).isoformat()}
    await db.products.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/products/{product_id}")
async def update_product(product_id: str, req: ProductCreate, user=Depends(get_current_user)):
    await db.products.update_one({"id": product_id}, {"$set": req.model_dump()})
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    return product

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, user=Depends(get_current_user)):
    await db.products.delete_one({"id": product_id})
    return {"message": "Producto eliminado"}

@api_router.get("/products/{product_id}/bot-config")
async def get_product_bot_config(product_id: str, user=Depends(get_current_user)):
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    # Merge stored config with defaults to ensure all fields are present
    defaults = {"personality": "", "key_benefits": "", "usage_info": "", "restrictions": "", "faqs": "", "sales_flow": ""}
    stored_config = product.get("bot_config", {})
    return {**defaults, **stored_config}

@api_router.put("/products/{product_id}/bot-config")
async def update_product_bot_config(product_id: str, config: dict, user=Depends(get_current_user)):
    allowed = {"personality", "key_benefits", "usage_info", "restrictions", "faqs", "sales_flow"}
    clean = {k: v for k, v in config.items() if k in allowed}
    await db.products.update_one({"id": product_id}, {"$set": {"bot_config": clean}})
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    return product.get("bot_config", {})

# ========== BOT TRAINING CENTER (Developer only) ==========

@api_router.get("/bot-training/global-config")
async def get_bot_global_config(user=Depends(get_current_user)):
    """Get global bot configuration. Developer only."""
    if user.get("role") != "developer":
        raise HTTPException(status_code=403, detail="Solo desarrolladores")
    config = await db.bot_training.find_one({"id": "global"}, {"_id": 0})
    defaults = {
        "id": "global",
        "bot_name": "Asesor Virtual Fakulti",
        "brand_name": "Fakulti",
        "tone": "Cercano, experto, humano, confiable. Ciencia + natural = Biotecnología.",
        "greeting_style": "Saluda con un emoji y pregunta el nombre del cliente.",
        "farewell_style": "Despídete cordialmente y recuerda que estás disponible.",
        "prohibited_phrases": "No prometer curas. No afirmar que reemplaza tratamientos médicos. No usar lenguaje vulgar.",
        "general_instructions": "",
        "max_emojis_per_message": 2,
        "max_lines_per_message": 6,
        "response_language": "Español (Ecuador)"
    }
    return {**defaults, **(config or {})}

@api_router.put("/bot-training/global-config")
async def update_bot_global_config(config: dict, user=Depends(get_current_user)):
    """Update global bot configuration. Developer only."""
    if user.get("role") != "developer":
        raise HTTPException(status_code=403, detail="Solo desarrolladores")
    allowed = {"bot_name", "brand_name", "tone", "greeting_style", "farewell_style", "prohibited_phrases", "general_instructions", "max_emojis_per_message", "max_lines_per_message", "response_language"}
    clean = {k: v for k, v in config.items() if k in allowed}
    clean["id"] = "global"
    await db.bot_training.update_one({"id": "global"}, {"$set": clean}, upsert=True)
    result = await db.bot_training.find_one({"id": "global"}, {"_id": 0})
    return result

@api_router.get("/bot-training/knowledge-base")
async def get_knowledge_base(user=Depends(get_current_user)):
    """Get FAQ/knowledge base entries. Developer only."""
    if user.get("role") != "developer":
        raise HTTPException(status_code=403, detail="Solo desarrolladores")
    entries = await db.knowledge_base.find({}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return entries

@api_router.post("/bot-training/knowledge-base")
async def add_knowledge_entry(body: dict, user=Depends(get_current_user)):
    """Add a knowledge base entry. Developer only."""
    if user.get("role") != "developer":
        raise HTTPException(status_code=403, detail="Solo desarrolladores")
    entry = {
        "id": str(uuid.uuid4()),
        "question": body.get("question", ""),
        "answer": body.get("answer", ""),
        "category": body.get("category", "general"),
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.knowledge_base.insert_one(entry)
    entry.pop("_id", None)
    return entry

@api_router.put("/bot-training/knowledge-base/{entry_id}")
async def update_knowledge_entry(entry_id: str, body: dict, user=Depends(get_current_user)):
    if user.get("role") != "developer":
        raise HTTPException(status_code=403, detail="Solo desarrolladores")
    update = {}
    for k in ("question", "answer", "category", "active"):
        if k in body:
            update[k] = body[k]
    if update:
        await db.knowledge_base.update_one({"id": entry_id}, {"$set": update})
    entry = await db.knowledge_base.find_one({"id": entry_id}, {"_id": 0})
    return entry

@api_router.delete("/bot-training/knowledge-base/{entry_id}")
async def delete_knowledge_entry(entry_id: str, user=Depends(get_current_user)):
    if user.get("role") != "developer":
        raise HTTPException(status_code=403, detail="Solo desarrolladores")
    await db.knowledge_base.delete_one({"id": entry_id})
    return {"message": "Entrada eliminada"}

@api_router.post("/bot-training/test")
async def test_bot_response(body: dict, user=Depends(get_current_user)):
    """Test the bot with a simulated conversation. Developer only. Fully isolated from real DB."""
    if user.get("role") != "developer":
        raise HTTPException(status_code=403, detail="Solo desarrolladores")
    from emergentintegrations.llm.chat import LlmChat, UserMessage

    test_message = body.get("message", "Hola")
    conversation_history = body.get("history", [])

    # Load products & global config for prompt building (read-only)
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    product_info = "\n".join([f"- {p['name']}: ${p['price']} - {p.get('description', '')}" for p in products])

    global_config = await db.bot_training.find_one({"id": "global"}, {"_id": 0}) or {}
    bot_name = global_config.get("bot_name", "Asesor Virtual Fakulti")
    brand_name = global_config.get("brand_name", "Fakulti")
    tone = global_config.get("tone", "Cercano, experto, humano, confiable. Ciencia + natural = Biotecnología.")
    greeting_style = global_config.get("greeting_style", "Saluda con un emoji y pregunta el nombre del cliente.")
    farewell_style = global_config.get("farewell_style", "Despídete cordialmente y recuerda que estás disponible.")
    prohibited = global_config.get("prohibited_phrases", "No prometer curas. No afirmar que reemplaza tratamientos médicos.")
    general_inst = global_config.get("general_instructions", "")
    max_emojis = global_config.get("max_emojis_per_message", 2)
    max_lines = global_config.get("max_lines_per_message", 6)

    kb_entries = await db.knowledge_base.find({"active": True}, {"_id": 0, "question": 1, "answer": 1}).to_list(50)
    kb_text = ""
    if kb_entries:
        kb_text = "\nBASE DE CONOCIMIENTO:\n" + "\n".join([f"P: {e['question']}\nR: {e['answer']}" for e in kb_entries])

    # Extract data from conversation history to avoid re-asking
    collected = []
    all_user_text = "\n".join([m["content"] for m in conversation_history if m.get("role") == "user"])
    all_user_text += "\n" + test_message
    import re as _re
    name_tags = _re.findall(r'\[LEAD_NAME:([^\]]+)\]', "\n".join([m["content"] for m in conversation_history if m.get("role") == "assistant"]))
    if name_tags:
        collected.append(f"Nombre: {name_tags[-1].strip()}")
    product_tags = _re.findall(r'\[UPDATE_LEAD:product_interest=([^\]]+)\]', "\n".join([m["content"] for m in conversation_history if m.get("role") == "assistant"]))
    if product_tags:
        collected.append(f"Producto de interés: {product_tags[-1].strip()}")
    city_tags = _re.findall(r'\[UPDATE_LEAD:city=([^\]]+)\]', "\n".join([m["content"] for m in conversation_history if m.get("role") == "assistant"]))
    if city_tags:
        collected.append(f"Ciudad: {city_tags[-1].strip()}")

    collected_text = ""
    if collected:
        collected_text = "\nDATOS YA PROPORCIONADOS POR EL CLIENTE (PROHIBIDO volver a solicitar):\n" + "\n".join(f"- {d}" for d in collected)

    recent = conversation_history[-10:]
    conv_summary = "\n".join([f"{'CLIENTE' if m['role']=='user' else 'BOT'}: {m['content'][:200]}" for m in recent])
    if conv_summary:
        collected_text += f"\n\nRESUMEN ULTIMOS MENSAJES:\n{conv_summary}"

    system_msg = f"""IDENTIDAD DEL AGENTE
Eres {bot_name}, el asesor virtual de la marca {brand_name} por WhatsApp.
Tu tono y estilo: {tone}
Habla como persona real, no como robot. Frases cortas y faciles de entender.
Puedes usar emojis de forma natural (máximo {max_emojis} por mensaje).
Despedida: {farewell_style}
{collected_text}
{f"INSTRUCCIONES ADICIONALES DEL DESARROLLADOR: {general_inst}" if general_inst else ""}

REGLA CRITICA - NO REPETIR PREGUNTAS
Lee TODA la conversación anterior antes de responder. Si el cliente YA proporcionó un dato en CUALQUIER mensaje anterior, NO lo pidas de nuevo.

TODOS LOS PRODUCTOS:
{product_info}

FLUJO DE CONVERSACION
1. Si no tienes el nombre, saluda y pregunta nombre.
2. Una vez tengas el nombre, saluda "Hola [nombre], mucho gusto" y pregunta que producto le interesa.
3. Cuando identifiques el producto de interes, incluye: [UPDATE_LEAD:product_interest=NombreProducto]

RESPUESTAS CORTAS: entre 1 y {max_lines} lineas.

PROHIBIDO
{prohibited}
- NO uses markdown, negritas, asteriscos ni formatos especiales. Solo texto plano.
{kb_text}

MODO TEST: Esta es una conversación de prueba del desarrollador. Responde normalmente como si fuera un cliente real.

EXTRACCION AUTOMATICA DE DATOS
Al final de CADA respuesta, incluye en lineas separadas si detectas datos nuevos:
- [LEAD_NAME:Nombre Apellido]
- [UPDATE_LEAD:city=Ciudad]
- [UPDATE_LEAD:email=correo@ejemplo.com]
- [UPDATE_LEAD:product_interest=NombreProducto]
- [STAGE:nuevo|interesado|en_negociacion|cliente_nuevo|perdido]"""

    llm_key = os.environ.get('EMERGENT_LLM_KEY')
    session_id = f"test_dev_{uuid.uuid4().hex[:8]}"
    chat = LlmChat(api_key=llm_key, session_id=session_id, system_message=system_msg)
    chat.with_model("openai", "gpt-5.2")

    # Feed previous conversation history
    for msg in conversation_history:
        chat.messages.append({"role": msg["role"], "content": msg["content"]})

    try:
        response = await chat.send_message(UserMessage(text=test_message))
        reply = response if isinstance(response, str) else str(response)
    except Exception as e:
        logger.error(f"Test bot GPT error: {e}")
        reply = "Error al obtener respuesta del bot. Verifica la configuración."

    # Clean tags from display reply
    clean_reply = re.sub(r'\[LEAD_NAME:[^\]]+\]', '', reply)
    clean_reply = re.sub(r'\[UPDATE_LEAD:\w+=[^\]]+\]', '', clean_reply)
    clean_reply = re.sub(r'\[STAGE:\w+\]', '', clean_reply).strip()

    return {"reply": clean_reply}

@api_router.get("/bot-training/admins")
async def get_admin_users_for_dev(user=Depends(get_current_user)):
    """Get admin users list. Developer only (for password reset)."""
    if user.get("role") != "developer":
        raise HTTPException(status_code=403, detail="Solo desarrolladores")
    admins = await db.admin_users.find({"role": "admin"}, {"_id": 0, "password_hash": 0}).to_list(50)
    return admins

# ========== GAME ROUTES ==========

@api_router.get("/games/config")
async def get_games_config(user=Depends(get_current_user)):
    configs = await db.games_config.find({}, {"_id": 0}).to_list(100)
    return configs

@api_router.post("/games/config")
async def create_game_config(req: GameConfigCreate, user=Depends(get_current_user)):
    doc = {"id": str(uuid.uuid4()), **req.model_dump(), "created_at": datetime.now(timezone.utc).isoformat()}
    await db.games_config.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/games/config/{config_id}")
async def update_game_config(config_id: str, req: GameConfigCreate, user=Depends(get_current_user)):
    await db.games_config.update_one({"id": config_id}, {"$set": req.model_dump()})
    config = await db.games_config.find_one({"id": config_id}, {"_id": 0})
    return config

@api_router.get("/games/public/{game_type}")
async def get_game_public(game_type: str):
    config = await db.games_config.find_one({"game_type": game_type, "active": True}, {"_id": 0})
    if not config:
        raise HTTPException(status_code=404, detail="Juego no disponible")
    safe_config = {"game_type": config["game_type"], "name": config["name"], "prizes": [{"name": p["name"], "color": p.get("color", "#A3E635")} for p in config["prizes"]]}
    return safe_config

@api_router.post("/games/play")
async def play_game(req: GamePlayRequest):
    config = await db.games_config.find_one({"game_type": req.game_type, "active": True}, {"_id": 0})
    if not config:
        raise HTTPException(status_code=404, detail="Juego no disponible")
    
    max_plays = config.get("max_plays_per_whatsapp", 1)
    plays_count = await db.game_plays.count_documents({"whatsapp": req.whatsapp, "game_type": req.game_type})
    if plays_count >= max_plays:
        raise HTTPException(status_code=400, detail="Ya has jugado el maximo permitido")
    
    prizes = config["prizes"]
    weights = [p.get("probability", 1) for p in prizes]
    total_weight = sum(weights)
    normalized = [w / total_weight for w in weights]
    selected_prize = secrets.SystemRandom().choices(prizes, weights=normalized, k=1)[0]
    
    play_doc = {
        "id": str(uuid.uuid4()),
        "whatsapp": req.whatsapp,
        "name": req.name,
        "game_type": req.game_type,
        "prize": selected_prize["name"],
        "prize_data": selected_prize,
        "converted": False,
        "played_at": datetime.now(timezone.utc).isoformat()
    }
    await db.game_plays.insert_one(play_doc)
    
    existing_lead = await find_lead_by_phone(req.whatsapp)
    if existing_lead:
        await db.leads.update_one(
            {"id": existing_lead["id"]},
            {"$set": {"game_used": req.game_type, "prize_obtained": selected_prize["name"], "last_interaction": datetime.now(timezone.utc).isoformat(), "whatsapp": normalize_phone_ec(req.whatsapp)}}
        )
    else:
        lead_doc = {
            "id": str(uuid.uuid4()),
            "name": req.name,
            "whatsapp": normalize_phone_ec(req.whatsapp),
            "city": req.city or "",
            "email": "",
            "product_interest": "",
            "source": "QR",
            "game_used": req.game_type,
            "prize_obtained": selected_prize["name"],
            "funnel_stage": "interesado",
            "status": "activo",
            "purchase_history": [],
            "coupon_used": None,
            "recompra_date": None,
            "notes": f"Registro via juego {req.game_type}",
            "last_interaction": datetime.now(timezone.utc).isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.leads.insert_one(lead_doc)
    
    prize_index = prizes.index(selected_prize)
    return {"prize": selected_prize["name"], "prize_index": prize_index, "coupon": selected_prize.get("coupon", ""), "message": selected_prize.get("message", f"Ganaste: {selected_prize['name']}")}

@api_router.get("/games/plays")
async def get_game_plays(game_type: Optional[str] = None, user=Depends(get_current_user)):
    query = {}
    if game_type:
        query["game_type"] = game_type
    plays = await db.game_plays.find(query, {"_id": 0}).sort("played_at", -1).limit(200).to_list(200)
    return plays

# ========== QUOTATION ROUTES ==========

@api_router.get("/quotations")
async def get_quotations(user=Depends(get_current_user)):
    quotations = await db.quotations.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return quotations

@api_router.post("/quotations")
async def create_quotation(req: QuotationCreate, user=Depends(get_current_user)):
    lead = await db.leads.find_one({"id": req.lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    
    subtotal = sum(item.get("price", 0) * item.get("quantity", 1) for item in req.items)
    tax = round(subtotal * 0.12, 2)
    total = round(subtotal + tax, 2)
    
    doc = {
        "id": str(uuid.uuid4()),
        "lead_id": req.lead_id,
        "lead_name": lead.get("name", ""),
        "lead_whatsapp": lead.get("whatsapp", ""),
        "items": req.items,
        "subtotal": round(subtotal, 2),
        "tax": tax,
        "total": total,
        "notes": req.notes,
        "status": "pendiente",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.quotations.insert_one(doc)
    doc.pop("_id", None)
    
    await db.leads.update_one(
        {"id": req.lead_id},
        {"$set": {"funnel_stage": "en_negociacion", "last_interaction": datetime.now(timezone.utc).isoformat()}}
    )
    return doc

@api_router.get("/quotations/{quotation_id}/pdf")
async def get_quotation_pdf(quotation_id: str):
    from fpdf import FPDF
    
    quotation = await db.quotations.find_one({"id": quotation_id}, {"_id": 0})
    if not quotation:
        raise HTTPException(status_code=404, detail="Cotizacion no encontrada")
    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 20)
    pdf.cell(0, 15, "FAKULTI LABORATORIOS", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 8, "La Ciencia de lo Natural", ln=True, align="C")
    pdf.ln(10)
    
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, f"Cotizacion #{quotation_id[:8]}", ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 7, f"Fecha: {quotation['created_at'][:10]}", ln=True)
    pdf.cell(0, 7, f"Cliente: {quotation['lead_name']}", ln=True)
    pdf.cell(0, 7, f"WhatsApp: {quotation['lead_whatsapp']}", ln=True)
    pdf.ln(10)
    
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(163, 230, 53)
    pdf.cell(80, 8, "Producto", border=1, fill=True)
    pdf.cell(25, 8, "Cant.", border=1, fill=True, align="C")
    pdf.cell(35, 8, "Precio Unit.", border=1, fill=True, align="C")
    pdf.cell(35, 8, "Subtotal", border=1, fill=True, align="C")
    pdf.ln()
    
    pdf.set_font("Helvetica", "", 10)
    for item in quotation["items"]:
        qty = item.get("quantity", 1)
        price = item.get("price", 0)
        pdf.cell(80, 7, item.get("name", ""), border=1)
        pdf.cell(25, 7, str(qty), border=1, align="C")
        pdf.cell(35, 7, f"${price:.2f}", border=1, align="C")
        pdf.cell(35, 7, f"${price * qty:.2f}", border=1, align="C")
        pdf.ln()
    
    pdf.ln(5)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(140, 8, "Subtotal:", align="R")
    pdf.cell(35, 8, f"${quotation['subtotal']:.2f}", align="C")
    pdf.ln()
    pdf.cell(140, 8, "IVA (12%):", align="R")
    pdf.cell(35, 8, f"${quotation['tax']:.2f}", align="C")
    pdf.ln()
    pdf.cell(140, 8, "TOTAL:", align="R")
    pdf.cell(35, 8, f"${quotation['total']:.2f}", align="C")
    
    if quotation.get("notes"):
        pdf.ln(15)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(0, 7, f"Notas: {quotation['notes']}", ln=True)
    
    buffer = io.BytesIO()
    pdf.output(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=cotizacion_{quotation_id[:8]}.pdf"})

# ========== LOYALTY ROUTES ==========

@api_router.get("/loyalty/sequences")
async def get_loyalty_sequences(user=Depends(get_current_user)):
    sequences = await db.loyalty_sequences.find({}, {"_id": 0}).to_list(100)
    return sequences

@api_router.post("/loyalty/sequences")
async def create_loyalty_sequence(req: LoyaltySequenceCreate, user=Depends(get_current_user)):
    doc = {"id": str(uuid.uuid4()), **req.model_dump(), "created_at": datetime.now(timezone.utc).isoformat()}
    await db.loyalty_sequences.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/loyalty/sequences/{sequence_id}")
async def update_loyalty_sequence(sequence_id: str, req: LoyaltySequenceCreate, user=Depends(get_current_user)):
    await db.loyalty_sequences.update_one({"id": sequence_id}, {"$set": req.model_dump()})
    seq = await db.loyalty_sequences.find_one({"id": sequence_id}, {"_id": 0})
    return seq

@api_router.delete("/loyalty/sequences/{sequence_id}")
async def delete_loyalty_sequence(sequence_id: str, user=Depends(get_current_user)):
    await db.loyalty_sequences.delete_one({"id": sequence_id})
    return {"message": "Secuencia eliminada"}

# ---- Loyalty Enrollments ----

@api_router.get("/loyalty/enrollments")
async def get_loyalty_enrollments(user=Depends(get_current_user)):
    enrollments = await db.loyalty_enrollments.find({}, {"_id": 0}).sort("enrolled_at", -1).to_list(200)
    return enrollments

@api_router.post("/loyalty/enroll")
async def enroll_lead_loyalty(lead_id: str = Query(...), sequence_id: str = Query(...), user=Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    seq = await db.loyalty_sequences.find_one({"id": sequence_id}, {"_id": 0})
    if not seq:
        raise HTTPException(status_code=404, detail="Secuencia no encontrada")
    existing = await db.loyalty_enrollments.find_one({"lead_id": lead_id, "sequence_id": sequence_id})
    if existing:
        raise HTTPException(status_code=400, detail="Lead ya esta inscrito en esta secuencia")
    now = datetime.now(timezone.utc)
    messages_schedule = []
    for msg in seq.get("messages", []):
        if msg.get("active", True):
            send_date = now + timedelta(days=msg.get("day", 1))
            messages_schedule.append({
                "day": msg["day"],
                "content": msg["content"],
                "scheduled_date": send_date.isoformat(),
                "status": "pendiente"
            })
    enrollment = {
        "id": str(uuid.uuid4()),
        "lead_id": lead_id,
        "lead_name": lead.get("name", ""),
        "lead_whatsapp": lead.get("whatsapp", ""),
        "sequence_id": sequence_id,
        "sequence_name": seq.get("product_name", ""),
        "messages": messages_schedule,
        "status": "activo",
        "enrolled_at": now.isoformat(),
        "completed_at": None
    }
    await db.loyalty_enrollments.insert_one(enrollment)
    enrollment.pop("_id", None)
    return enrollment

@api_router.delete("/loyalty/enrollments/{enrollment_id}")
async def delete_enrollment(enrollment_id: str, user=Depends(get_current_user)):
    result = await db.loyalty_enrollments.delete_one({"id": enrollment_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Inscripcion no encontrada")
    return {"message": "Inscripcion eliminada"}

@api_router.post("/loyalty/process")
async def process_loyalty_messages(user=Depends(get_current_user)):
    """Process all pending loyalty messages that are due."""
    now = datetime.now(timezone.utc)
    enrollments = await db.loyalty_enrollments.find({"status": "activo"}, {"_id": 0}).to_list(500)
    sent_count = 0
    for enrollment in enrollments:
        updated = False
        all_done = True
        for msg in enrollment.get("messages", []):
            if msg["status"] == "pendiente":
                scheduled = datetime.fromisoformat(msg["scheduled_date"].replace("Z", "+00:00")) if "Z" in msg.get("scheduled_date", "") else datetime.fromisoformat(msg["scheduled_date"])
                if scheduled.tzinfo is None:
                    scheduled = scheduled.replace(tzinfo=timezone.utc)
                if now >= scheduled:
                    msg["status"] = "enviado"
                    msg["sent_at"] = now.isoformat()
                    sent_count += 1
                    updated = True
                    logger.info(f"Loyalty msg sent to {enrollment['lead_name']} ({enrollment['lead_whatsapp']}): Day {msg['day']}")
                else:
                    all_done = False
            elif msg["status"] != "enviado":
                all_done = False
        if updated:
            update_data = {"messages": enrollment["messages"]}
            if all_done:
                update_data["status"] = "completado"
                update_data["completed_at"] = now.isoformat()
            await db.loyalty_enrollments.update_one({"id": enrollment["id"]}, {"$set": update_data})
    return {"processed": sent_count, "message": f"{sent_count} mensajes procesados"}

@api_router.get("/loyalty/metrics")
async def get_loyalty_metrics(user=Depends(get_current_user)):
    """Dashboard de métricas de fidelización basado en datos reales del CRM."""
    
    # 1. Lead stats
    all_leads = await db.leads.find({}, {"_id": 0, "id": 1, "name": 1, "funnel_stage": 1, "product_interest": 1, "channel": 1, "source": 1, "created_at": 1, "last_interaction": 1, "purchase_history": 1}).to_list(1000)
    total_leads = len(all_leads)
    
    # Funnel distribution
    stage_map = {"nuevo": "Contacto inicial", "interesado": "Chat", "en_negociacion": "En Negociación", "cliente_nuevo": "Leads ganados", "cliente_activo": "Cartera activa", "perdido": "Perdido"}
    funnel_data = {}
    for lead in all_leads:
        stage = lead.get("funnel_stage", "nuevo")
        label = stage_map.get(stage, stage)
        funnel_data[label] = funnel_data.get(label, 0) + 1
    funnel_distribution = [{"stage": k, "count": v} for k, v in funnel_data.items()]
    
    # Product interest distribution
    product_interest = {}
    for lead in all_leads:
        p = lead.get("product_interest", "")
        if p:
            product_interest[p] = product_interest.get(p, 0) + 1
    product_interest_data = [{"product": k, "leads": v} for k, v in sorted(product_interest.items(), key=lambda x: x[1], reverse=True)]
    
    # Source distribution
    source_data = {}
    for lead in all_leads:
        s = lead.get("source", "") or "Sin fuente"
        source_data[s] = source_data.get(s, 0) + 1
    source_distribution = [{"source": k, "count": v} for k, v in source_data.items()]
    
    # 2. Conversation stats
    total_sessions = await db.chat_sessions_meta.count_documents({})
    total_messages = await db.chat_messages.count_documents({})
    bot_messages = await db.chat_messages.count_documents({"role": "assistant"})
    user_messages = await db.chat_messages.count_documents({"role": "user"})
    campaign_messages = await db.chat_messages.count_documents({"source": "campaign"})
    
    # Active leads (with recent interaction)
    converted = len([lead for lead in all_leads if lead.get("funnel_stage") in ["cliente_nuevo", "cliente_activo"]])
    in_progress = len([lead for lead in all_leads if lead.get("funnel_stage") in ["interesado", "en_negociacion"]])
    lost = len([lead for lead in all_leads if lead.get("funnel_stage") == "perdido"])
    conversion_rate = round((converted / total_leads * 100) if total_leads > 0 else 0, 1)
    
    # 3. Campaign stats
    campaigns = await db.campaigns.find({}, {"_id": 0, "name": 1, "sent_count": 1, "failed_count": 1, "status": 1, "target_count": 1}).to_list(50)
    total_campaign_sends = sum(c.get("sent_count", 0) for c in campaigns)
    
    # 4. Purchase-based metrics (when purchase_history exists)
    clients_with_purchases = [lead for lead in all_leads if lead.get("purchase_history") and len(lead["purchase_history"]) > 0]
    total_clients = len(clients_with_purchases)
    repeat_buyers = [c for c in clients_with_purchases if len(c.get("purchase_history", [])) > 1]
    total_revenue = sum(sum(p.get("price", 0) for p in c.get("purchase_history", [])) for c in clients_with_purchases)
    total_purchases = sum(len(c.get("purchase_history", [])) for c in clients_with_purchases)
    avg_order_value = round(total_revenue / total_purchases, 2) if total_purchases > 0 else 0
    
    # Product revenue from purchases
    product_revenue = {}
    for lead in clients_with_purchases:
        for purchase in lead.get("purchase_history", []):
            pname = purchase.get("product_name", "Desconocido")
            product_revenue[pname] = product_revenue.get(pname, 0) + purchase.get("price", 0)
    product_revenue_data = [{"product": k, "revenue": round(v, 2)} for k, v in sorted(product_revenue.items(), key=lambda x: x[1], reverse=True)]
    
    # Top buyers
    top_buyers_list = []
    for b in sorted(clients_with_purchases, key=lambda c: sum(p.get("price", 0) for p in c.get("purchase_history", [])), reverse=True)[:10]:
        top_buyers_list.append({
            "name": b.get("name", ""),
            "purchases": len(b.get("purchase_history", [])),
            "total_spent": round(sum(p.get("price", 0) for p in b.get("purchase_history", [])), 2),
            "stage": stage_map.get(b.get("funnel_stage", ""), b.get("funnel_stage", ""))
        })
    
    # 5. Loyalty sequence stats
    all_enrollments = await db.loyalty_enrollments.find({}, {"_id": 0}).to_list(500)
    total_enrollments = len(all_enrollments)
    active_enrollments = len([e for e in all_enrollments if e.get("status") == "activo"])
    completed_enrollments = len([e for e in all_enrollments if e.get("status") == "completado"])
    total_seq_msgs = sum(len(e.get("messages", [])) for e in all_enrollments)
    sent_seq_msgs = sum(len([m for m in e.get("messages", []) if m.get("status") == "enviado"]) for e in all_enrollments)
    
    seq_stats = {}
    for e in all_enrollments:
        sid = e.get("sequence_name", "Desconocida")
        if sid not in seq_stats:
            seq_stats[sid] = {"name": sid, "enrollments": 0, "completed": 0, "msgs_sent": 0, "msgs_total": 0}
        seq_stats[sid]["enrollments"] += 1
        if e.get("status") == "completado":
            seq_stats[sid]["completed"] += 1
        for m in e.get("messages", []):
            seq_stats[sid]["msgs_total"] += 1
            if m.get("status") == "enviado":
                seq_stats[sid]["msgs_sent"] += 1
    sequence_effectiveness = []
    for s in seq_stats.values():
        s["completion_rate"] = round((s["completed"] / s["enrollments"] * 100) if s["enrollments"] > 0 else 0, 1)
        s["delivery_rate"] = round((s["msgs_sent"] / s["msgs_total"] * 100) if s["msgs_total"] > 0 else 0, 1)
        sequence_effectiveness.append(s)
    
    return {
        "summary": {
            "total_leads": total_leads,
            "converted": converted,
            "in_progress": in_progress,
            "lost": lost,
            "conversion_rate": conversion_rate,
            "total_sessions": total_sessions,
            "total_messages": total_messages,
            "bot_messages": bot_messages,
            "user_messages": user_messages,
            "campaign_messages": campaign_messages,
            "total_campaign_sends": total_campaign_sends,
            "total_clients": total_clients,
            "repeat_buyers": len(repeat_buyers),
            "total_revenue": round(total_revenue, 2),
            "avg_order_value": avg_order_value,
            "active_clients": converted,
            "lost_clients": lost
        },
        "funnel_distribution": funnel_distribution,
        "product_interest": product_interest_data,
        "source_distribution": source_distribution,
        "product_revenue": product_revenue_data,
        "top_buyers": top_buyers_list,
        "loyalty": {
            "total_enrollments": total_enrollments,
            "active_enrollments": active_enrollments,
            "completed_enrollments": completed_enrollments,
            "total_messages": total_seq_msgs,
            "sent_messages": sent_seq_msgs,
            "delivery_rate": round((sent_seq_msgs / total_seq_msgs * 100) if total_seq_msgs > 0 else 0, 1)
        },
        "sequence_effectiveness": sequence_effectiveness,
        "campaigns": [{"name": c.get("name", ""), "sent": c.get("sent_count", 0), "failed": c.get("failed_count", 0), "status": c.get("status", "")} for c in campaigns]
    }

# ========== CHAT HELPERS ==========

async def _resolve_chat_lead(session_id: str, lead_id: str = None):
    """Resolve lead from session metadata or lead_id."""
    session_meta = await db.chat_sessions_meta.find_one({"session_id": session_id}, {"_id": 0})
    lead = None
    has_name = False
    if session_meta and session_meta.get("lead_id"):
        lead = await db.leads.find_one({"id": session_meta["lead_id"]}, {"_id": 0})
        if lead:
            has_name = True
    elif lead_id:
        lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
        if lead:
            has_name = True
            await db.chat_sessions_meta.update_one(
                {"session_id": session_id},
                {"$set": {"session_id": session_id, "lead_id": lead["id"], "lead_name": lead.get("name", "")}},
                upsert=True
            )
    return lead, has_name


async def _parse_ai_response(assistant_content: str, has_name: bool, lead_id_for_session, session_id: str):
    """Parse AI response for lead creation, updates, and stage classification."""
    name_match = re.search(r'\[LEAD_NAME:([^\]]+)\]', assistant_content)
    if name_match and not has_name:
        detected_name = name_match.group(1).strip()
        assistant_content = re.sub(r'\[LEAD_NAME:[^\]]+\]', '', assistant_content).strip()
        new_lead = {
            "id": str(uuid.uuid4()), "name": detected_name, "whatsapp": "", "city": "", "email": "",
            "product_interest": "", "source": "Chat IA", "game_used": None, "prize_obtained": None,
            "funnel_stage": "nuevo", "status": "activo", "purchase_history": [], "coupon_used": None,
            "recompra_date": None, "notes": "Registrado vía Chat IA",
            "last_interaction": datetime.now(timezone.utc).isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.leads.insert_one(new_lead)
        lead_id_for_session = new_lead["id"]
        await db.chat_sessions_meta.update_one(
            {"session_id": session_id},
            {"$set": {"session_id": session_id, "lead_id": new_lead["id"], "lead_name": detected_name}},
            upsert=True
        )
        await db.chat_messages.update_many(
            {"session_id": session_id, "lead_id": None}, {"$set": {"lead_id": new_lead["id"]}}
        )
        logger.info(f"New lead created from chat: {detected_name} -> {new_lead['id']}")

    update_matches = re.findall(r'\[UPDATE_LEAD:(\w+)=([^\]]+)\]', assistant_content)
    if update_matches and lead_id_for_session:
        update_fields = {}
        for field, value in update_matches:
            if field in {"whatsapp", "city", "product_interest", "email"}:
                update_fields[field] = value.strip()
        if update_fields:
            update_fields["last_interaction"] = datetime.now(timezone.utc).isoformat()
            await db.leads.update_one({"id": lead_id_for_session}, {"$set": update_fields})
            logger.info(f"Lead {lead_id_for_session} updated via chat: {update_fields}")
        assistant_content = re.sub(r'\[UPDATE_LEAD:\w+=[^\]]+\]', '', assistant_content).strip()

    stage_match = re.search(r'\[STAGE:(\w+)\]', assistant_content)
    if stage_match:
        new_stage = stage_match.group(1).strip()
        assistant_content = re.sub(r'\[STAGE:\w+\]', '', assistant_content).strip()
        if new_stage in FUNNEL_STAGES and lead_id_for_session:
            await db.leads.update_one(
                {"id": lead_id_for_session},
                {"$set": {"funnel_stage": new_stage, "last_interaction": datetime.now(timezone.utc).isoformat()}}
            )

    return assistant_content, lead_id_for_session


# ========== CHAT ROUTES ==========

@api_router.post("/chat/message")
async def send_chat_message(req: ChatMessageRequest, user=Depends(get_current_user)):
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    product_info = "\n".join([f"- {p['name']}: ${p['price']} - {p.get('description', '')}" for p in products])
    
    lead, has_name = await _resolve_chat_lead(req.session_id, req.lead_id)
    lead_name = lead.get("name", "") if lead else ""
    
    # Check for pending quotation
    pending_quote = ""
    if lead:
        quote = await db.quotations.find_one({"lead_id": lead["id"], "status": "pendiente"}, {"_id": 0})
        if quote:
            pending_quote = f"\nEste cliente tiene una cotización pendiente por ${quote['total']:.2f}. Pregunta si desea continuar con ella."
    
    # Use product-specific bot if product interest is identified
    product_interest = lead.get("product_interest", "") if lead else ""
    product_specific_prompt = None
    if product_interest:
        product_specific_prompt = await build_product_bot_prompt(product_interest, products, lead or {})
    
    if product_specific_prompt:
        system_msg = product_specific_prompt + (f"\n{pending_quote}" if pending_quote else "")
    else:
        # Build general context-aware system prompt
        name_instruction = ""
        if not has_name:
            name_instruction = """
IMPORTANTE - REGISTRO DE LEAD:
- Este es un lead NUEVO. Saluda cordialmente y pregunta su nombre completo.
- Una vez que proporcione su nombre, confirma diciendo su nombre y pregunta su numero de WhatsApp.
- Despues pregunta su ciudad y que producto le interesa.
- Cuando el usuario diga su nombre, incluye al FINAL de tu respuesta (en una linea separada): [LEAD_NAME:Nombre Apellido]
- Solo incluye [LEAD_NAME:] cuando el usuario efectivamente diga su nombre.
- Recopila los datos uno por uno de forma natural, no todos de golpe."""
        else:
            missing_fields = []
            if lead and not lead.get("whatsapp"):
                missing_fields.append("numero de WhatsApp")
            if lead and not lead.get("city"):
                missing_fields.append("ciudad")
            if lead and not lead.get("product_interest"):
                missing_fields.append("que producto le interesa")
            missing_instruction = ""
            if missing_fields:
                missing_instruction = f"\nDATOS FALTANTES DEL CLIENTE: Necesitas preguntarle su {', '.join(missing_fields)}. Hazlo de forma natural durante la conversación."
                missing_instruction += "\nCuando el cliente proporcione datos, incluye al final: [UPDATE_LEAD:campo=valor] donde campo puede ser: whatsapp, city, product_interest"
            name_instruction = f"\nEl cliente se llama {lead_name}. Usa su nombre de forma natural en la conversación.{missing_instruction}"

        system_msg = f"""Eres el Asesor Virtual Oficial de Fakulti Laboratorios (marca Fakulti).
Tu función: Atender leads, calificar, cotizar y cerrar venta.
{name_instruction}
{pending_quote}

PRODUCTOS DISPONIBLES:
{product_info}

REGLAS:
- Nunca sonar robot. Sé amigable y profesional.
- Nunca hacer promesas médicas.
- Nunca afirmar que cura enfermedades.
- Nunca recomendar reemplazar tratamiento médico.
- Bombro es Bone Broth Hidrolizado, producto único en Ecuador.
- Responde siempre en español.
- Sé conciso pero útil.
- Si el usuario pide precio, proporciona la información.
- Si pide comprar, indica los pasos.

DETECCIÓN DE PRODUCTO:
Cuando identifiques qué producto le interesa al cliente, incluye: [UPDATE_LEAD:product_interest=NombreProducto]

CLASIFICACIÓN AUTOMÁTICA:
Al final de CADA respuesta, incluye en una línea separada la etapa del lead:
[STAGE:nuevo] - Primer contacto, aún no muestra interés específico
[STAGE:interesado] - Pregunta por productos, precios o beneficios
[STAGE:en_negociacion] - Solicita cotización, forma de pago, envío o stock
[STAGE:cliente_nuevo] - Confirma compra
[STAGE:perdido] - Dice que no le interesa o rechaza explícitamente

Incluye SIEMPRE el tag [STAGE:] al final."""

    history = await db.chat_messages.find(
        {"session_id": req.session_id}, {"_id": 0}
    ).sort("timestamp", 1).limit(20).to_list(20)
    
    llm_key = os.environ.get('EMERGENT_LLM_KEY')
    chat = LlmChat(api_key=llm_key, session_id=req.session_id, system_message=system_msg)
    chat.with_model("openai", "gpt-5.2")
    
    for msg in history:
        if msg["role"] == "user":
            chat.messages.append({"role": "user", "content": msg["content"]})
        else:
            chat.messages.append({"role": "assistant", "content": msg["content"]})
    
    user_msg_doc = {
        "id": str(uuid.uuid4()),
        "session_id": req.session_id,
        "lead_id": req.lead_id or (lead["id"] if lead else None),
        "role": "user",
        "content": req.message,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    await db.chat_messages.insert_one(user_msg_doc)
    
    try:
        response = await chat.send_message(UserMessage(text=req.message))
        assistant_content = response if isinstance(response, str) else str(response)
    except Exception as e:
        logger.error(f"Chat error: {e}")
        assistant_content = "Disculpa, tuve un problema tecnico. Puedes repetir tu consulta?"
    
    # Parse AI response for lead management
    lead_id_for_session = req.lead_id or (lead["id"] if lead else None)
    assistant_content, lead_id_for_session = await _parse_ai_response(
        assistant_content, has_name, lead_id_for_session, req.session_id
    )
    
    assistant_msg_doc = {
        "id": str(uuid.uuid4()),
        "session_id": req.session_id,
        "lead_id": lead_id_for_session,
        "role": "assistant",
        "content": assistant_content,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    await db.chat_messages.insert_one(assistant_msg_doc)
    
    # Get updated lead info for response
    lead_info = None
    if lead_id_for_session:
        lead_doc = await db.leads.find_one({"id": lead_id_for_session}, {"_id": 0})
        if lead_doc:
            lead_info = {"id": lead_doc["id"], "name": lead_doc["name"], "funnel_stage": lead_doc["funnel_stage"]}
    
    return {"response": assistant_content, "session_id": req.session_id, "lead": lead_info}

@api_router.delete("/chat/messages/{message_id}")
async def delete_chat_message(message_id: str, user=Depends(get_current_user)):
    result = await db.chat_messages.delete_one({"id": message_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Mensaje no encontrado")
    return {"message": "Mensaje eliminado"}

@api_router.delete("/chat/sessions/{session_id}")
async def delete_chat_session(session_id: str, user=Depends(get_current_user)):
    result = await db.chat_messages.delete_many({"session_id": session_id})
    await db.chat_sessions_meta.delete_one({"session_id": session_id})
    return {"message": f"Conversacion eliminada ({result.deleted_count} mensajes)"}

@api_router.get("/chat/history/{session_id}")
async def get_chat_history(session_id: str, user=Depends(get_current_user)):
    messages = await db.chat_messages.find({"session_id": session_id}, {"_id": 0}).sort("timestamp", 1).to_list(100)
    return messages

@api_router.get("/chat/sessions")
async def get_chat_sessions(user=Depends(get_current_user)):
    pipeline = [
        {"$group": {"_id": "$session_id", "lead_id": {"$first": "$lead_id"}, "last_message": {"$last": "$content"}, "timestamp": {"$last": "$timestamp"}, "count": {"$sum": 1}}},
        {"$sort": {"timestamp": -1}},
        {"$limit": 50}
    ]
    sessions = await db.chat_messages.aggregate(pipeline).to_list(50)
    user_role = user.get("role", "admin")
    user_id = user["id"]
    result = []
    for s in sessions:
        meta = await db.chat_sessions_meta.find_one({"session_id": s["_id"]}, {"_id": 0})
        lead_name = meta.get("lead_name", "") if meta else ""
        source = meta.get("source", "chat_ia") if meta else "chat_ia"
        lead_phone = ""
        lead_channel = ""
        bot_paused = False
        has_alert = False
        assigned_advisor = ""
        needs_advisor = False
        if source == "whatsapp" and s.get("lead_id"):
            lead_doc = await db.leads.find_one({"id": s["lead_id"]}, {"_id": 0, "whatsapp": 1, "name": 1, "channel": 1, "bot_paused": 1, "assigned_advisor": 1, "needs_advisor": 1, "funnel_stage": 1})
            if lead_doc:
                lead_phone = lead_doc.get("whatsapp", "")
                lead_channel = lead_doc.get("channel", "")
                bot_paused = lead_doc.get("bot_paused", False)
                assigned_advisor = lead_doc.get("assigned_advisor", "")
                needs_advisor = lead_doc.get("needs_advisor", False) and not assigned_advisor
                if not lead_name:
                    lead_name = lead_doc.get("name", "")
            alert = await db.handover_alerts.find_one({"lead_id": s["lead_id"], "status": "pending"}, {"_id": 0})
            has_alert = alert is not None
        # Role-based filtering: advisors only see their assigned conversations
        if user_role == "advisor" and assigned_advisor != user_id:
            continue
        result.append({
            "session_id": s["_id"], "lead_id": s.get("lead_id"), "lead_name": lead_name,
            "last_message": s["last_message"], "timestamp": s["timestamp"], "message_count": s["count"],
            "source": source, "lead_phone": lead_phone, "lead_channel": lead_channel, "bot_paused": bot_paused, "has_alert": has_alert,
            "assigned_advisor": assigned_advisor, "needs_advisor": needs_advisor
        })
    return result

@api_router.get("/chat/lead-session/{lead_id}")
async def get_or_create_lead_session(lead_id: str, user=Depends(get_current_user)):
    """Find existing chat session for a lead, or return a new session_id to use."""
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    if lead.get("assigned_advisor"):
        advisor = await db.admin_users.find_one({"id": lead["assigned_advisor"], "role": "advisor"}, {"_id": 0, "name": 1})
        lead["_advisor_name"] = advisor["name"] if advisor else ""
    meta = await db.chat_sessions_meta.find_one({"lead_id": lead_id}, {"_id": 0})
    if meta:
        msgs = await db.chat_messages.find({"session_id": meta["session_id"]}, {"_id": 0}).sort("timestamp", 1).to_list(100)
        return {"session_id": meta["session_id"], "lead": lead, "messages": msgs, "is_new": False}
    new_sid = f"lead_{lead_id}_{int(datetime.now(timezone.utc).timestamp())}"
    await db.chat_sessions_meta.insert_one({"session_id": new_sid, "lead_id": lead_id, "lead_name": lead.get("name", "")})
    return {"session_id": new_sid, "lead": lead, "messages": [], "is_new": True}

# ========== WHATSAPP MONITORING ==========

@api_router.get("/chat/whatsapp-stats")
async def get_whatsapp_stats(user=Depends(get_current_user)):
    """Get WhatsApp monitoring stats."""
    # Active conversations (had activity in last 24h)
    cutoff = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
    active_sessions = await db.chat_sessions_meta.count_documents({"source": "whatsapp", "last_activity": {"$gte": cutoff}})
    total_wa_sessions = await db.chat_sessions_meta.count_documents({"source": "whatsapp"})
    
    # Average response time from recent WhatsApp messages
    pipeline = [
        {"$match": {"source": "whatsapp", "role": "assistant", "response_time_ms": {"$exists": True}}},
        {"$sort": {"timestamp": -1}},
        {"$limit": 50},
        {"$group": {"_id": None, "avg_response_ms": {"$avg": "$response_time_ms"}, "min_response_ms": {"$min": "$response_time_ms"}, "max_response_ms": {"$max": "$response_time_ms"}}}
    ]
    stats = await db.chat_messages.aggregate(pipeline).to_list(1)
    avg_ms = int(stats[0]["avg_response_ms"]) if stats else 0
    
    # Pending alerts
    pending_alerts = await db.handover_alerts.count_documents({"status": "pending"})
    
    # Messages today
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0).isoformat()
    messages_today = await db.chat_messages.count_documents({"source": "whatsapp", "timestamp": {"$gte": today_start}})
    
    # Delivery stats
    delivered = await db.chat_messages.count_documents({"source": "whatsapp", "role": "assistant", "delivered": True})
    failed = await db.chat_messages.count_documents({"source": "whatsapp", "role": "assistant", "delivered": False})
    
    return {
        "active_conversations_24h": active_sessions,
        "total_conversations": total_wa_sessions,
        "avg_response_time_ms": avg_ms,
        "pending_alerts": pending_alerts,
        "messages_today": messages_today,
        "delivered": delivered,
        "failed": failed
    }

@api_router.get("/chat/alerts")
async def get_handover_alerts(user=Depends(get_current_user)):
    """Get human handover alerts with lead context."""
    alerts = await db.handover_alerts.find({}, {"_id": 0}).sort("created_at", -1).to_list(50)
    # Enrich alerts with lead data
    for a in alerts:
        if a.get("lead_id"):
            lead = await db.leads.find_one({"id": a["lead_id"]}, {"_id": 0, "name": 1, "whatsapp": 1, "product_interest": 1, "channel": 1, "city": 1, "funnel_stage": 1, "bot_paused": 1})
            if lead:
                a["lead_name"] = lead.get("name", a.get("lead_name", ""))
                a["lead_product"] = lead.get("product_interest", "")
                a["lead_channel"] = lead.get("channel", "")
                a["lead_city"] = lead.get("city", "")
                a["lead_stage"] = lead.get("funnel_stage", "")
                a["bot_paused"] = lead.get("bot_paused", False)
    return alerts

@api_router.put("/chat/alerts/{alert_id}/resolve")
async def resolve_alert(alert_id: str, user=Depends(get_current_user)):
    """Mark a handover alert as resolved."""
    result = await db.handover_alerts.update_one(
        {"id": alert_id},
        {"$set": {"status": "resolved", "resolved_at": datetime.now(timezone.utc).isoformat(), "resolved_by": "admin"}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Alerta no encontrada")
    return {"message": "Alerta resuelta"}

@api_router.put("/leads/{lead_id}/pause-bot")
async def pause_bot_for_lead(lead_id: str, user=Depends(get_current_user)):
    """Pause bot automation for a specific lead (human takes over)."""
    result = await db.leads.update_one(
        {"id": lead_id},
        {"$set": {"bot_paused": True, "bot_paused_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    logger.info(f"Bot paused for lead {lead_id} - human agent taking over")
    return {"message": "Bot pausado. El agente humano tiene el control."}

@api_router.put("/leads/{lead_id}/resume-bot")
async def resume_bot_for_lead(lead_id: str, user=Depends(get_current_user)):
    """Resume bot automation for a specific lead."""
    result = await db.leads.update_one(
        {"id": lead_id},
        {"$set": {"bot_paused": False}, "$unset": {"bot_paused_at": ""}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    logger.info(f"Bot resumed for lead {lead_id}")
    return {"message": "Bot reactivado para este lead."}

@api_router.post("/chat/whatsapp-reply")
async def crm_whatsapp_reply(req: CRMWhatsAppReply, user=Depends(get_current_user)):
    """Send a message to a lead via WhatsApp from the CRM."""
    lead = await db.leads.find_one({"id": req.lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    phone = lead.get("whatsapp", "")
    if not phone:
        raise HTTPException(status_code=400, detail="Lead no tiene numero de WhatsApp")
    
    sent = await send_whatsapp_message(phone, req.message)
    if not sent:
        raise HTTPException(status_code=500, detail="Error al enviar mensaje")
    
    # Store in chat history
    session_id = f"wa_{phone}"
    now = datetime.now(timezone.utc).isoformat()
    await db.chat_messages.insert_one({
        "id": str(uuid.uuid4()), "session_id": session_id, "lead_id": req.lead_id,
        "role": "assistant", "content": req.message, "timestamp": now,
        "source": "whatsapp", "sent_by": "crm_agent", "delivered": sent
    })
    await db.chat_sessions_meta.update_one(
        {"session_id": session_id},
        {"$set": {"last_activity": now}},
        upsert=True
    )
    return {"message": "Mensaje enviado", "delivered": sent}

# ========== BULK ROUTES ==========

@api_router.post("/bulk/upload")
async def bulk_upload(file: UploadFile = File(...), user=Depends(get_current_user)):
    import openpyxl
    
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    headers = [cell.value for cell in ws[1] if cell.value]
    header_map = {}
    for i, h in enumerate(headers):
        h_lower = h.lower().strip()
        if "nombre" in h_lower or "name" in h_lower:
            header_map["name"] = i
        elif "whatsapp" in h_lower or "telefono" in h_lower or "phone" in h_lower:
            header_map["whatsapp"] = i
        elif "ciudad" in h_lower or "city" in h_lower:
            header_map["city"] = i
        elif "producto" in h_lower or "product" in h_lower:
            header_map["product_interest"] = i
        elif "fecha" in h_lower or "date" in h_lower:
            header_map["purchase_date"] = i
        elif "email" in h_lower:
            header_map["email"] = i
        elif "fuente" in h_lower or "source" in h_lower:
            header_map["source"] = i
        elif "temporada" in h_lower or "season" in h_lower:
            header_map["season"] = i
        elif "canal" in h_lower or "channel" in h_lower:
            header_map["channel"] = i
    
    created = 0
    updated = 0
    errors = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            name = str(row[header_map.get("name", 0)] or "").strip()
            whatsapp_raw = str(row[header_map.get("whatsapp", 1)] or "").strip()
            if not name or not whatsapp_raw:
                continue
            whatsapp = normalize_phone_ec(whatsapp_raw)
            
            city = str(row[header_map.get("city", 2)] or "").strip() if "city" in header_map else ""
            product = str(row[header_map.get("product_interest", 3)] or "").strip() if "product_interest" in header_map else ""
            email = str(row[header_map.get("email", -1)] or "").strip() if "email" in header_map else ""
            source_val = str(row[header_map.get("source", -1)] or "").strip() if "source" in header_map else "Carga masiva"
            season_val = str(row[header_map.get("season", -1)] or "").strip() if "season" in header_map else ""
            channel_val = str(row[header_map.get("channel", -1)] or "").strip() if "channel" in header_map else ""
            
            has_purchase = "purchase_date" in header_map and row[header_map["purchase_date"]]
            stage = "cliente_nuevo" if has_purchase else "nuevo"
            
            existing = await find_lead_by_phone(whatsapp)
            if existing:
                update_data = {"name": name, "whatsapp": whatsapp, "last_interaction": datetime.now(timezone.utc).isoformat()}
                if city:
                    update_data["city"] = city
                if product:
                    update_data["product_interest"] = product
                if has_purchase:
                    update_data["funnel_stage"] = stage
                await db.leads.update_one({"id": existing["id"]}, {"$set": update_data})
                updated += 1
            else:
                lead_doc = {
                    "id": str(uuid.uuid4()),
                    "name": name,
                    "whatsapp": whatsapp,
                    "city": city,
                    "email": email,
                    "product_interest": product,
                    "source": source_val,
                    "season": season_val,
                    "channel": channel_val,
                    "game_used": None,
                    "prize_obtained": None,
                    "funnel_stage": stage,
                    "status": "activo",
                    "purchase_history": [],
                    "coupon_used": None,
                    "recompra_date": None,
                    "notes": "",
                    "last_interaction": datetime.now(timezone.utc).isoformat(),
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.leads.insert_one(lead_doc)
                created += 1
        except Exception as e:
            logger.error(f"Bulk upload row error: {e}")
            errors += 1
    
    return {"created": created, "updated": updated, "errors": errors, "total_processed": created + updated + errors}

@api_router.get("/bulk/download")
async def bulk_download(
    download_type: str = Query("all"),
    stage: Optional[str] = None,
    product: Optional[str] = None,
    user=Depends(get_current_user)
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    query = {}
    if download_type == "stage" and stage:
        query["funnel_stage"] = stage
    elif download_type == "product" and product:
        query["product_interest"] = {"$regex": product, "$options": "i"}
    elif download_type == "fibeca":
        query["source"] = "Fibeca"
    elif download_type == "game":
        query["game_used"] = {"$ne": None}
    elif download_type == "recompra":
        query["funnel_stage"] = {"$in": ["cliente_nuevo", "cliente_activo"]}
    
    leads = await db.leads.find(query, {"_id": 0}).to_list(10000)
    all_leads = await db.leads.find({}, {"_id": 0}).to_list(10000)
    products_list = await db.products.find({}, {"_id": 0}).to_list(100)
    
    wb = openpyxl.Workbook()
    
    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A6B3C", end_color="1A6B3C", fill_type="solid")
    header_fill2 = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_fill3 = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB")
    )
    center_align = Alignment(horizontal="center", vertical="center")
    
    def style_header(ws, row, fill):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = fill
            cell.alignment = center_align
            cell.border = thin_border
    
    def style_data(ws, start_row):
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
    
    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 35)
    
    # ===== SHEET 1: Base de Datos de Leads =====
    ws1 = wb.active
    ws1.title = "Base de Datos"
    
    stage_labels = {"nuevo": "Contacto inicial", "interesado": "Chat", "en_negociacion": "En Negociación", "cliente_nuevo": "Leads ganados", "cliente_activo": "Cartera activa", "perdido": "Perdido"}
    
    headers = ["#", "Nombre", "WhatsApp", "Email", "Ciudad", "Producto de Interés", "Fuente", "Etapa del Embudo", "Estado", "Juego Usado", "Premio Obtenido", "Cupón", "Última Interacción", "Fecha de Registro"]
    ws1.append(headers)
    style_header(ws1, 1, header_fill)
    
    for i, lead in enumerate(leads, 1):
        ws1.append([
            i,
            lead.get("name", ""),
            lead.get("whatsapp", ""),
            lead.get("email", ""),
            lead.get("city", ""),
            lead.get("product_interest", ""),
            lead.get("source", ""),
            stage_labels.get(lead.get("funnel_stage", ""), lead.get("funnel_stage", "")),
            lead.get("status", ""),
            lead.get("game_used", "") or "",
            lead.get("prize_obtained", "") or "",
            lead.get("coupon_used", "") or "",
            (lead.get("last_interaction", "") or "")[:19],
            (lead.get("created_at", "") or "")[:19]
        ])
    style_data(ws1, 2)
    auto_width(ws1)
    ws1.freeze_panes = "B2"
    ws1.auto_filter.ref = f"A1:N{len(leads)+1}"
    
    # ===== SHEET 2: Resumen por Etapa del Embudo =====
    ws2 = wb.create_sheet("Embudo de Ventas")
    ws2.append(["Etapa del Embudo", "Cantidad de Leads", "Porcentaje", "Tasa de Conversión"])
    style_header(ws2, 1, header_fill2)
    
    stage_order = ["nuevo", "interesado", "en_negociacion", "cliente_nuevo", "cliente_activo", "perdido"]
    total_leads = len(all_leads)
    for s in stage_order:
        count = sum(1 for lead in all_leads if lead.get("funnel_stage") == s)
        pct = f"{(count/total_leads*100):.1f}%" if total_leads > 0 else "0%"
        conv = ""
        if s == "cliente_nuevo" or s == "cliente_activo":
            conv = f"{(count/total_leads*100):.1f}%" if total_leads > 0 else "0%"
        ws2.append([stage_labels.get(s, s), count, pct, conv])
    ws2.append([])
    ws2.append(["TOTAL LEADS", total_leads, "100%", ""])
    ws2[ws2.max_row][0].font = Font(bold=True)
    ws2[ws2.max_row][1].font = Font(bold=True)
    style_data(ws2, 2)
    auto_width(ws2)
    
    # ===== SHEET 3: Leads por Fuente =====
    ws3 = wb.create_sheet("Fuentes de Tráfico")
    ws3.append(["Fuente", "Total Leads", "Porcentaje", "Clientes Convertidos", "Tasa de Conversión"])
    style_header(ws3, 1, header_fill3)
    
    sources = {}
    for lead in all_leads:
        src = lead.get("source", "Desconocido") or "Desconocido"
        if src not in sources:
            sources[src] = {"total": 0, "clientes": 0}
        sources[src]["total"] += 1
        if lead.get("funnel_stage") in ["cliente_nuevo", "cliente_activo"]:
            sources[src]["clientes"] += 1
    
    for src, data in sorted(sources.items(), key=lambda x: x[1]["total"], reverse=True):
        pct = f"{(data['total']/total_leads*100):.1f}%" if total_leads > 0 else "0%"
        conv = f"{(data['clientes']/data['total']*100):.1f}%" if data["total"] > 0 else "0%"
        ws3.append([src, data["total"], pct, data["clientes"], conv])
    style_data(ws3, 2)
    auto_width(ws3)
    
    # ===== SHEET 4: Productos =====
    ws4 = wb.create_sheet("Catálogo de Productos")
    ws4.append(["Producto", "Código", "Precio", "Precio Original", "Stock", "Categoría", "Leads Interesados"])
    style_header(ws4, 1, header_fill)
    
    for p in products_list:
        interested = sum(1 for lead in all_leads if p["name"].lower() in (lead.get("product_interest", "") or "").lower())
        ws4.append([p["name"], p.get("code", ""), f"${p['price']}", f"${p.get('original_price', '')}", p.get("stock", ""), p.get("category", ""), interested])
    style_data(ws4, 2)
    auto_width(ws4)
    
    # ===== SHEET 5: Leads por Ciudad =====
    ws5 = wb.create_sheet("Leads por Ciudad")
    ws5.append(["Ciudad", "Total Leads", "Porcentaje", "Clientes"])
    style_header(ws5, 1, header_fill2)
    
    cities = {}
    for lead in all_leads:
        city = lead.get("city", "Sin ciudad") or "Sin ciudad"
        if city not in cities:
            cities[city] = {"total": 0, "clientes": 0}
        cities[city]["total"] += 1
        if lead.get("funnel_stage") in ["cliente_nuevo", "cliente_activo"]:
            cities[city]["clientes"] += 1
    
    for city, data in sorted(cities.items(), key=lambda x: x[1]["total"], reverse=True):
        pct = f"{(data['total']/total_leads*100):.1f}%" if total_leads > 0 else "0%"
        ws5.append([city, data["total"], pct, data["clientes"]])
    style_data(ws5, 2)
    auto_width(ws5)
    
    # ===== SHEET 6: Leads por Período =====
    ws6 = wb.create_sheet("Leads por Período")
    ws6.append(["Período", "Nuevos Leads", "Leads Interesados", "En Negociación", "Clientes Nuevos", "Perdidos", "Total"])
    style_header(ws6, 1, header_fill3)
    
    months = {}
    for lead in all_leads:
        created = lead.get("created_at", "")
        if created:
            month_key = created[:7]
            if month_key not in months:
                months[month_key] = {"nuevo": 0, "interesado": 0, "en_negociacion": 0, "cliente_nuevo": 0, "perdido": 0, "total": 0}
            stage_val = lead.get("funnel_stage", "nuevo")
            if stage_val in months[month_key]:
                months[month_key][stage_val] += 1
            months[month_key]["total"] += 1
    
    for period in sorted(months.keys()):
        d = months[period]
        ws6.append([period, d["nuevo"], d["interesado"], d["en_negociacion"], d["cliente_nuevo"], d["perdido"], d["total"]])
    style_data(ws6, 2)
    auto_width(ws6)
    
    # ===== SHEET 7: Resumen Ejecutivo =====
    ws7 = wb.create_sheet("Resumen Ejecutivo")
    ws7.sheet_properties.tabColor = "FFD700"
    ws7.move_range("A1", rows=0, cols=0)
    
    title_font = Font(name="Calibri", bold=True, size=16, color="1A6B3C")
    subtitle_font = Font(name="Calibri", bold=True, size=12, color="333333")
    
    ws7.append(["REPORTE EJECUTIVO - FAKULTI LABORATORIOS"])
    ws7["A1"].font = title_font
    ws7.merge_cells("A1:D1")
    ws7.append([f"Generado: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} UTC"])
    ws7["A2"].font = Font(italic=True, color="666666")
    ws7.append([])
    
    ws7.append(["KPI", "Valor"])
    ws7["A4"].font = subtitle_font
    ws7["B4"].font = subtitle_font
    
    clientes = sum(1 for lead in all_leads if lead.get("funnel_stage") in ["cliente_nuevo", "cliente_activo"])
    interesados = sum(1 for lead in all_leads if lead.get("funnel_stage") == "interesado")
    negociacion = sum(1 for lead in all_leads if lead.get("funnel_stage") == "en_negociacion")
    perdidos = sum(1 for lead in all_leads if lead.get("funnel_stage") == "perdido")
    
    kpis = [
        ("Total Leads", total_leads),
        ("Clientes Activos", clientes),
        ("Leads Interesados", interesados),
        ("En Negociación", negociacion),
        ("Leads Perdidos", perdidos),
        ("Tasa de Conversión", f"{(clientes/total_leads*100):.1f}%" if total_leads > 0 else "0%"),
        ("Fuente Principal", max(sources.items(), key=lambda x: x[1]["total"])[0] if sources else "N/A"),
        ("Ciudad Principal", max(cities.items(), key=lambda x: x[1]["total"])[0] if cities else "N/A"),
    ]
    for kpi, val in kpis:
        ws7.append([kpi, val])
    
    style_data(ws7, 4)
    auto_width(ws7)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"fakulti_reporte_{download_type}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})

# ========== WHATSAPP CLOUD API ==========

WHATSAPP_API_URL = "https://graph.facebook.com/v25.0"

HANDOVER_KEYWORDS = ["agente", "humano", "persona real", "hablar con alguien", "asesor real", "no quiero bot", "operador", "representante", "persona de verdad", "quiero hablar con una persona", "atención humana"]

# Timeout for bot to resolve (seconds) - if bot interaction exceeds this without progress, trigger alert
BOT_TIMEOUT_SECONDS = 60

async def get_whatsapp_config():
    config = await db.whatsapp_config.find_one({"id": "main"}, {"_id": 0})
    return config or {"id": "main", "phone_number_id": "", "access_token": "", "verify_token": "fakulti-whatsapp-verify-token", "business_name": "Fakulti Laboratorios"}

async def send_whatsapp_message(to_phone: str, text: str):
    """Send a message via WhatsApp Cloud API."""
    config = await get_whatsapp_config()
    if not config.get("phone_number_id") or not config.get("access_token"):
        logger.warning("WhatsApp not configured - message not sent")
        return False
    international_phone = phone_to_international(to_phone)
    url = f"{WHATSAPP_API_URL}/{config['phone_number_id']}/messages"
    headers = {"Authorization": f"Bearer {config['access_token']}", "Content-Type": "application/json"}
    payload = {"messaging_product": "whatsapp", "to": international_phone, "type": "text", "text": {"body": text}}
    try:
        async with httpx.AsyncClient() as client_http:
            resp = await client_http.post(url, json=payload, headers=headers, timeout=15)
            if resp.status_code in (200, 201):
                logger.info(f"WhatsApp message sent to {to_phone}")
                return True
            else:
                logger.error(f"WhatsApp send error: {resp.status_code} {resp.text}")
                return False
    except Exception as e:
        logger.error(f"WhatsApp send exception: {e}")
        return False

async def send_whatsapp_image(to_phone: str, image_url: str, caption: str = ""):
    """Send an image with optional caption via WhatsApp Cloud API."""
    config = await get_whatsapp_config()
    if not config.get("phone_number_id") or not config.get("access_token"):
        logger.warning("WhatsApp not configured - image not sent")
        return False
    international_phone = phone_to_international(to_phone)
    url = f"{WHATSAPP_API_URL}/{config['phone_number_id']}/messages"
    headers = {"Authorization": f"Bearer {config['access_token']}", "Content-Type": "application/json"}
    image_payload = {"link": image_url}
    if caption:
        image_payload["caption"] = caption
    payload = {"messaging_product": "whatsapp", "to": international_phone, "type": "image", "image": image_payload}
    try:
        async with httpx.AsyncClient() as client_http:
            resp = await client_http.post(url, json=payload, headers=headers, timeout=15)
            if resp.status_code in (200, 201):
                logger.info(f"WhatsApp image sent to {to_phone}")
                return True
            else:
                logger.error(f"WhatsApp image send error: {resp.status_code} {resp.text}")
                return False
    except Exception as e:
        logger.error(f"WhatsApp image send exception: {e}")
        return False

UPLOADS_DIR = Path(__file__).parent / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)

@api_router.post("/upload-image")
async def upload_image(file: UploadFile = File(...), user=Depends(get_current_user)):
    """Upload an image and return its public URL."""
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Solo se permiten imágenes")
    ext = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = UPLOADS_DIR / filename
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Imagen muy grande (máx 5MB)")
    with open(filepath, "wb") as f:
        f.write(content)
    return {"url": f"/api/uploads/{filename}", "filename": filename}

@api_router.get("/uploads/{filename}")
async def serve_upload(filename: str):
    filepath = UPLOADS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    return FileResponse(filepath)

async def build_product_bot_prompt(product_name: str, all_products: list, lead_data: dict) -> str:
    """Build a product-specific bot prompt. Only includes info about the target product."""
    target = None
    for p in all_products:
        if product_name.lower() in p["name"].lower() or p["name"].lower() in product_name.lower():
            target = p
            break
    
    if not target:
        return None
    
    # Get bot config from product or use defaults
    bot_cfg = target.get("bot_config", {})
    personality = bot_cfg.get("personality", "experto amigable en el producto")
    key_benefits = bot_cfg.get("key_benefits", target.get("description", ""))
    usage_info = bot_cfg.get("usage_info", "Consultar con un asesor para instrucciones de uso.")
    restrictions = bot_cfg.get("restrictions", "No hacer promesas medicas. No afirmar que cura enfermedades.")
    faqs = bot_cfg.get("faqs", "")
    sales_flow = bot_cfg.get("sales_flow", "")
    
    lead_name = lead_data.get("name", "")
    lead_city = lead_data.get("city", "")
    lead_email = lead_data.get("email", "")
    collected_data_text = lead_data.get("_collected_data_text", "")
    
    # Build data context
    data_context = ""
    if lead_name:
        data_context = f"\nEl cliente se llama {lead_name}."
        if lead_city:
            data_context += f" Ciudad: {lead_city}."
        if lead_email:
            data_context += f" Email: {lead_email}."
        if lead_data.get("ci_ruc"):
            data_context += f" CI/RUC: {lead_data['ci_ruc']}."
        if lead_data.get("address"):
            data_context += f" Dirección: {lead_data['address']}."
    
    missing_fields = []
    if not lead_name:
        missing_fields.append("nombre y apellido")
    if not lead_city:
        missing_fields.append("ciudad")
    if not lead_email:
        missing_fields.append("email")
    missing_instruction = ""
    if missing_fields:
        missing_instruction = f"\nDATOS FALTANTES: Recopila de forma natural: {', '.join(missing_fields)}."
    
    first_contact = "\nEste es un lead NUEVO. Saluda y pregunta su nombre." if not lead_name else ""
    
    # Other products list (just names and prices for redirection)
    other_products = [f"- {p['name']}: ${p['price']}" for p in all_products if p["id"] != target["id"]]
    other_products_text = "\n".join(other_products) if other_products else "No hay otros productos."
    
    # If there's a detailed sales_flow, use it as the primary prompt structure
    if sales_flow:
        return f"""IDENTIDAD DEL AGENTE
Eres el asesor virtual especializado en {target['name']} de la marca Fakulti por WhatsApp.
Personalidad: {personality}
Tu estilo: Cercano, experto, humano, confiable (no robotico). Ciencia + natural = Biotecnologia.
Habla como persona real, NO como robot. Frases cortas. Emojis moderados (1-2 por mensaje).
{first_contact}
{data_context}
{collected_data_text}

REGLA CRITICA - NO REPETIR PREGUNTAS
Lee TODA la conversación anterior. Si el cliente YA proporcionó un dato (nombre, teléfono, ciudad, dirección, cédula, cantidad, etc.) en CUALQUIER mensaje anterior, NO lo vuelvas a pedir. Usa la información de la conversación. Si un dato ya fue mencionado, simplemente avanza al siguiente paso del flujo.

TU PRODUCTO: {target['name']}
Codigo: {target.get('code', '')}
Precio oferta: ${target['price']}
{f"Precio normal: ${target.get('original_price', '')}" if target.get('original_price') else ""}

=== FLUJO DE VENTAS COMPLETO ===
{sales_flow}
=== FIN DEL FLUJO ===

REGLA CRITICA - PRODUCTO UNICO
Solo puedes hablar sobre {target['name']}. NO mezcles informacion de otros productos.
Si el cliente pregunta por otro producto, responde:
"Claro, también tenemos otros productos. Te puedo conectar con información de ese producto."
Y lista brevemente:
{other_products_text}
Luego vuelve a tu producto principal.

RESTRICCIONES GENERALES
{restrictions}
- NO uses markdown, negritas, asteriscos ni formatos especiales. Solo texto plano y emojis.
- Si piden hablar con un humano, responde que un asesor se comunicará pronto.
- Respuestas CORTAS y CLARAS (máximo 4-6 líneas por mensaje). NO envíes bloques largos.
- Siempre lleva la conversación hacia el cierre de venta.
- Prioriza beneficios + resultado sobre información técnica.
- NUNCA repitas una pregunta que el cliente ya respondió en la conversación.

EXTRACCIÓN AUTOMÁTICA DE DATOS
Al final de CADA respuesta, incluye en líneas separadas:
- Si detectas nombre: [LEAD_NAME:Nombre Apellido]
- Si detectas ciudad: [UPDATE_LEAD:city=Ciudad]
- Si detectas email: [UPDATE_LEAD:email=correo@ejemplo.com]
- Si detectas CI/RUC: [UPDATE_LEAD:ci_ruc=valor]
- Si detectas dirección: [UPDATE_LEAD:address=dirección completa]
- Clasifica la etapa:
  [STAGE:nuevo] - Primer contacto
  [STAGE:interesado] - Pregunta por producto, precios o beneficios
  [STAGE:en_negociacion] - Solicita compra, pago, envío, pide info de precio
  [STAGE:cliente_nuevo] - Confirma compra, da datos de facturación
  [STAGE:perdido] - Rechaza explícitamente
Incluye SIEMPRE [STAGE:] al final."""
    
    # Fallback: original simple prompt for products without sales_flow
    return f"""IDENTIDAD DEL AGENTE
Eres el asesor virtual especializado en {target['name']} de la marca Fakulti por WhatsApp.
Personalidad: {personality}
Tu estilo: natural, cercano, humano, profesional, claro, breve.
Habla como persona real, no como robot. Frases cortas. Máximo 1-2 emojis por mensaje.
{first_contact}
{data_context}
{collected_data_text}
{missing_instruction}

REGLA CRITICA - NO REPETIR PREGUNTAS
Lee TODA la conversación anterior. Si el cliente YA proporcionó un dato en CUALQUIER mensaje anterior, NO lo pidas de nuevo. Avanza al siguiente paso.

TU PRODUCTO: {target['name']}
Código: {target.get('code', '')}
Precio: ${target['price']}
{f"Precio original: ${target.get('original_price', '')}" if target.get('original_price') else ""}
Descripción: {target.get('description', '')}
Beneficios clave: {key_benefits}
Cómo se usa: {usage_info}
{f"Preguntas frecuentes: {faqs}" if faqs else ""}

REGLA CRÍTICA - PRODUCTO ÚNICO
Solo puedes hablar sobre {target['name']}. NO mezcles información de otros productos.
Si el cliente pregunta por otro producto, responde:
"Claro, también tenemos otros productos. Te puedo conectar con información de ese producto. ¿Quieres que te cuente sobre alguno de estos?"
Y lista brevemente los otros productos disponibles:
{other_products_text}

Luego vuelve a tu producto principal: {target['name']}.

FLUJO
1. Si no tienes nombre, saluda y pregunta nombre.
2. Con nombre: "Hola [nombre], me alegra que te interese {target['name']}. Cuéntame, ¿ya conocías este producto?"
3. Adapta la explicación según las dudas del cliente.
4. Guía hacia compra sin presionar.

RESTRICCIONES
{restrictions}
- NO uses markdown, negritas, asteriscos ni formatos especiales. Solo texto plano.
- Si piden hablar con un humano, responde que un asesor se comunicará pronto.

EXTRACCIÓN AUTOMÁTICA DE DATOS
Al final de CADA respuesta, incluye en líneas separadas:
- Si detectas nombre: [LEAD_NAME:Nombre Apellido]
- Si detectas ciudad: [UPDATE_LEAD:city=Ciudad]
- Si detectas email: [UPDATE_LEAD:email=correo@ejemplo.com]
- Clasifica la etapa:
  [STAGE:nuevo] - Primer contacto
  [STAGE:interesado] - Pregunta por producto, precios o beneficios
  [STAGE:en_negociacion] - Solicita compra, pago, envío
  [STAGE:cliente_nuevo] - Confirma compra
  [STAGE:perdido] - Rechaza explícitamente
Incluye SIEMPRE [STAGE:] al final."""


async def process_whatsapp_incoming(phone: str, message_text: str):
    """Process an incoming WhatsApp message through GPT-5.2 AI bot."""
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    
    phone = normalize_phone_ec(phone)
    existing_lead = await find_lead_by_phone(phone)
    
    # If found with different format, normalize the stored phone
    if existing_lead and existing_lead.get("whatsapp") != phone:
        await db.leads.update_one({"id": existing_lead["id"]}, {"$set": {"whatsapp": phone}})
        existing_lead["whatsapp"] = phone
    
    is_new = existing_lead is None
    lead_id = existing_lead["id"] if existing_lead else None
    lead_name = existing_lead.get("name", "") if existing_lead else ""
    
    # CHECK: If bot is paused for this lead, don't auto-respond (human agent has control)
    if existing_lead and existing_lead.get("bot_paused"):
        # Still store the message in chat history for the agent to see
        session_id = f"wa_{phone}"
        now = datetime.now(timezone.utc).isoformat()
        await db.chat_messages.insert_one({"id": str(uuid.uuid4()), "session_id": session_id, "lead_id": lead_id, "role": "user", "content": message_text, "timestamp": now, "source": "whatsapp"})
        await db.chat_sessions_meta.update_one(
            {"session_id": session_id},
            {"$set": {"session_id": session_id, "lead_id": lead_id, "lead_name": lead_name, "source": "whatsapp", "last_activity": now}},
            upsert=True
        )
        await db.leads.update_one({"id": lead_id}, {"$set": {"last_interaction": now}})
        logger.info(f"Bot paused for lead {lead_id} ({phone}) - message stored, no auto-reply")
        return "[BOT_PAUSED] Mensaje recibido. Un asesor humano tiene el control.", lead_id
    
    # Create lead if new
    if is_new:
        new_lead = {
            "id": str(uuid.uuid4()), "name": "", "whatsapp": phone,
            "city": "", "email": "", "product_interest": "", "source": "WhatsApp",
            "season": "", "channel": "WhatsApp",
            "game_used": None, "prize_obtained": None, "funnel_stage": "nuevo", "status": "activo",
            "purchase_history": [], "coupon_used": None, "recompra_date": None,
            "notes": "Registrado vía WhatsApp",
            "last_interaction": datetime.now(timezone.utc).isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.leads.insert_one(new_lead)
        lead_id = new_lead["id"]
        existing_lead = new_lead
        # Auto-detect channel/source from first message (QR campaigns & intents)
        await detect_channel_from_message(message_text, lead_id)
        # Refresh lead data after potential update
        existing_lead = await db.leads.find_one({"id": lead_id}, {"_id": 0}) or existing_lead
    elif not existing_lead.get("channel") or existing_lead.get("channel") == "WhatsApp":
        # Check on subsequent messages too if channel not yet identified
        history_count = await db.chat_messages.count_documents({"session_id": f"wa_{phone}"})
        if history_count <= 2:
            detected = await detect_channel_from_message(message_text, lead_id)
            if detected:
                existing_lead = await db.leads.find_one({"id": lead_id}, {"_id": 0}) or existing_lead
    
    # Build product catalog
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    product_info = "\n".join([f"- {p['name']}: ${p['price']} - {p.get('description', '')}" for p in products])
    
    # ===== LOAD CONVERSATION HISTORY & BUILD DATA SUMMARY =====
    session_id = f"wa_{phone}"
    history = await db.chat_messages.find(
        {"session_id": session_id}, {"_id": 0}
    ).sort("timestamp", 1).to_list(50)
    
    all_user_text = "\n".join([m["content"] for m in history if m.get("role") == "user"])
    
    conversation_data = []
    if existing_lead.get("name"):
        conversation_data.append(f"Nombre: {existing_lead['name']}")
    if existing_lead.get("city"):
        conversation_data.append(f"Ciudad: {existing_lead['city']}")
    if existing_lead.get("email"):
        conversation_data.append(f"Email: {existing_lead['email']}")
    if existing_lead.get("product_interest"):
        conversation_data.append(f"Producto de interés: {existing_lead['product_interest']}")
    conversation_data.append(f"WhatsApp: {phone}")
    
    import re as _re
    ci_matches = _re.findall(r'(?:cedula|ci|ruc|cédula)[:\s]*(\d{10,13})', all_user_text, _re.IGNORECASE)
    if not ci_matches:
        ci_matches = _re.findall(r'\b(\d{10})\b', all_user_text)
    if ci_matches:
        ci_val = ci_matches[-1]
        conversation_data.append(f"CI/RUC mencionado: {ci_val}")
        if lead_id and not existing_lead.get("ci_ruc"):
            await db.leads.update_one({"id": lead_id}, {"$set": {"ci_ruc": ci_val}})
    
    phone_matches = _re.findall(r'(?:telefono|número|celular|contacto|cel)[:\s]*(09\d{8}|593\d{9})', all_user_text, _re.IGNORECASE)
    if not phone_matches:
        phone_matches = _re.findall(r'\b(09\d{8})\b', all_user_text)
    if phone_matches:
        conversation_data.append(f"Teléfono contacto: {phone_matches[-1]}")
        if lead_id and not existing_lead.get("whatsapp"):
            await db.leads.update_one({"id": lead_id}, {"$set": {"whatsapp": phone_matches[-1]}})
    
    addr_matches = _re.findall(r'(?:dirección|direccion|domicilio|entrega)[:\s]*(.{10,80})', all_user_text, _re.IGNORECASE)
    if addr_matches:
        conversation_data.append(f"Dirección: {addr_matches[-1].strip()}")
        if lead_id and not existing_lead.get("address"):
            await db.leads.update_one({"id": lead_id}, {"$set": {"address": addr_matches[-1].strip()}})
    list_addr = _re.findall(r'1\.\s*(.+(?:sauces|cdla|ciudadela|calle|mz|villa|manzana).+)', all_user_text, _re.IGNORECASE)
    if list_addr and not addr_matches:
        conversation_data.append(f"Dirección: {list_addr[-1].strip()}")
        if lead_id and not existing_lead.get("address"):
            await db.leads.update_one({"id": lead_id}, {"$set": {"address": list_addr[-1].strip()}})
    
    recent_msgs = history[-10:] if len(history) > 10 else history
    conversation_summary = "\n".join([f"{'CLIENTE' if m['role']=='user' else 'BOT'}: {m['content'][:200]}" for m in recent_msgs])
    
    collected_data_text = ""
    if conversation_data:
        collected_data_text = "\n\nDATOS YA PROPORCIONADOS POR EL CLIENTE (PROHIBIDO volver a solicitar):\n" + "\n".join(f"- {d}" for d in conversation_data)
    if conversation_summary:
        collected_data_text += f"\n\nRESUMEN ULTIMOS MENSAJES:\n{conversation_summary}"
    
    # Repetition detection & escalation
    bot_messages = [m["content"] for m in history if m.get("role") == "assistant"]
    repetition_keywords = ["dirección", "direccion", "cédula", "cedula", "contacto", "teléfono", "telefono", "número", "numero", "nombre completo"]
    ask_counts = {}
    for bm in bot_messages[-6:]:
        bm_lower = bm.lower()
        for kw in repetition_keywords:
            if kw in bm_lower and "?" in bm:
                ask_counts[kw] = ask_counts.get(kw, 0) + 1
    repeated_topics = [kw for kw, count in ask_counts.items() if count >= 2]
    if repeated_topics:
        advisor_id = existing_lead.get("assigned_advisor", "")
        await db.chat_alerts.insert_one({
            "id": str(uuid.uuid4()), "session_id": session_id, "lead_id": lead_id,
            "lead_name": existing_lead.get("name", phone), "type": "bot_confused",
            "message": f"Bot repite preguntas sobre: {', '.join(repeated_topics)}. Se requiere intervención.",
            "resolved": False, "created_at": datetime.now(timezone.utc).isoformat()
        })
        if advisor_id:
            await db.advisor_notifications.insert_one({
                "id": str(uuid.uuid4()), "advisor_id": advisor_id, "type": "bot_escalation",
                "title": f"Bot necesita ayuda con {existing_lead.get('name', phone)}",
                "message": f"Bot repite preguntas sobre {', '.join(repeated_topics)}.",
                "lead_id": lead_id, "session_id": session_id, "read": False,
                "created_at": datetime.now(timezone.utc).isoformat()
            })
        logger.warning(f"Bot escalation: repetition for lead {lead_id} on: {repeated_topics}")
        collected_data_text += f"\n\nALERTA: Cliente YA proporcionó {', '.join(repeated_topics)}. NO pidas estos datos. Si no puedes continuar, di: 'Un momento, te transfiero con un asesor.'"
    
    existing_lead["_collected_data_text"] = collected_data_text
    
    # ===== BUILD SYSTEM PROMPT =====
    
    # Check if lead has a product interest -> use product-specific bot
    product_interest = existing_lead.get("product_interest", "")
    product_specific_prompt = None
    if product_interest:
        product_specific_prompt = await build_product_bot_prompt(product_interest, products, existing_lead)
    
    if product_specific_prompt:
        system_msg = product_specific_prompt
    else:
        # General router bot - uses global config from Bot Training Center
        global_config = await db.bot_training.find_one({"id": "global"}, {"_id": 0}) or {}
        bot_name = global_config.get("bot_name", "Asesor Virtual Fakulti")
        brand_name = global_config.get("brand_name", "Fakulti")
        tone = global_config.get("tone", "Cercano, experto, humano, confiable. Ciencia + natural = Biotecnología.")
        greeting_style = global_config.get("greeting_style", "Saluda con un emoji y pregunta el nombre del cliente.")
        farewell_style = global_config.get("farewell_style", "Despídete cordialmente y recuerda que estás disponible.")
        prohibited = global_config.get("prohibited_phrases", "No prometer curas. No afirmar que reemplaza tratamientos médicos.")
        general_inst = global_config.get("general_instructions", "")
        max_emojis = global_config.get("max_emojis_per_message", 2)
        max_lines = global_config.get("max_lines_per_message", 6)
        
        # Load knowledge base FAQs
        kb_entries = await db.knowledge_base.find({"active": True}, {"_id": 0, "question": 1, "answer": 1}).to_list(50)
        kb_text = ""
        if kb_entries:
            kb_text = "\n\nBASE DE CONOCIMIENTO (usa estas respuestas cuando aplique):\n" + "\n".join(
                [f"P: {e['question']}\nR: {e['answer']}" for e in kb_entries]
            )
        
        # Build context about what data we already have and what's missing
        missing_fields = []
        if not lead_name:
            missing_fields.append("nombre y apellido")
        if not existing_lead.get("city"):
            missing_fields.append("ciudad")
        if not existing_lead.get("email"):
            missing_fields.append("email")
        if not existing_lead.get("product_interest"):
            missing_fields.append("producto de interes")
        
        data_context = ""
        if lead_name:
            data_context = f"\nEl cliente se llama {lead_name}."
            if existing_lead.get("city"):
                data_context += f" Ciudad: {existing_lead['city']}."
            if existing_lead.get("email"):
                data_context += f" Email: {existing_lead['email']}."
            if existing_lead.get("product_interest"):
                data_context += f" Interesado en: {existing_lead['product_interest']}."
        
        missing_instruction = ""
        if missing_fields:
            missing_instruction = f"\nDATOS QUE AUN FALTAN POR RECOPILAR: {', '.join(missing_fields)}. Recopilalos de forma natural durante la conversacion, uno a la vez, sin parecer formulario."
        
        first_contact = f"\nEste es un lead NUEVO. {greeting_style}" if not lead_name else ""
        
        system_msg = f"""IDENTIDAD DEL AGENTE
Eres {bot_name}, el asesor virtual de la marca {brand_name} por WhatsApp.
Representas los productos desarrollados por {brand_name} Laboratorios.
Tu tono y estilo: {tone}
Habla como persona real, no como robot. Frases cortas y faciles de entender.
Puedes usar emojis de forma natural (máximo {max_emojis} por mensaje).
Despedida: {farewell_style}
{first_contact}
{data_context}
{collected_data_text}
{missing_instruction}
{f"INSTRUCCIONES ADICIONALES DEL DESARROLLADOR: {general_inst}" if general_inst else ""}

REGLA CRITICA - NO REPETIR PREGUNTAS
Lee TODA la conversación anterior antes de responder. Si el cliente YA proporcionó un dato (nombre, teléfono, ciudad, dirección, cédula, etc.) en CUALQUIER mensaje anterior, NO lo pidas de nuevo. Usa la información que ya tienes. Si necesitas confirmar un dato, hazlo UNA sola vez.

TODOS LOS PRODUCTOS:
{product_info}

FLUJO DE CONVERSACION
1. Si no tienes el nombre, saluda y pregunta nombre.
2. Una vez tengas el nombre, saluda "Hola [nombre], mucho gusto" y pregunta que producto le interesa.
3. Cuando identifiques el producto de interes, incluye: [UPDATE_LEAD:product_interest=NombreProducto]
4. Una vez detectado el producto, enfocate en ese producto especificamente.

DETECCION DE PRODUCTO - MUY IMPORTANTE
Tu objetivo principal es identificar que producto le interesa al cliente.
Cuando el cliente mencione o muestre interes en un producto especifico, incluye:
[UPDATE_LEAD:product_interest=NombreExactoDelProducto]
Esto activara el bot especializado en ese producto para las siguientes interacciones.

COMO RESPONDER
Evita: "Gracias por su consulta", "Procedo a brindarle la informacion"
Usa: "Claro, te cuento", "Buena pregunta", "Mira, te explico rapido"

RESPUESTAS CORTAS: entre 1 y {max_lines} lineas.

PROHIBIDO
{prohibited}
- NO uses markdown, negritas, asteriscos ni formatos especiales. Solo texto plano.
- Si piden hablar con un humano, responde que un asesor se comunicara pronto.
- NUNCA repitas una pregunta que ya fue respondida en la conversación.
{kb_text}

EXTRACCION AUTOMATICA DE DATOS
Al final de CADA respuesta, incluye en lineas separadas:
- Si detectas nombre: [LEAD_NAME:Nombre Apellido]
- Si detectas ciudad: [UPDATE_LEAD:city=Ciudad]
- Si detectas email: [UPDATE_LEAD:email=correo@ejemplo.com]
- Si detectas producto de interes: [UPDATE_LEAD:product_interest=NombreProducto]
- Clasifica la etapa:
  [STAGE:nuevo] - Primer contacto
  [STAGE:interesado] - Pregunta por productos, precios o beneficios
  [STAGE:en_negociacion] - Solicita compra, pago, envio, cotizacion
  [STAGE:cliente_nuevo] - Confirma compra
  [STAGE:perdido] - Rechaza explicitamente
Incluye SIEMPRE [STAGE:] al final."""

    llm_key = os.environ.get('EMERGENT_LLM_KEY')
    chat = LlmChat(api_key=llm_key, session_id=session_id, system_message=system_msg)
    chat.with_model("openai", "gpt-5.2")
    
    for msg in history:
        if msg["role"] == "user":
            chat.messages.append({"role": "user", "content": msg["content"]})
        else:
            chat.messages.append({"role": "assistant", "content": msg["content"]})
    
    try:
        response = await chat.send_message(UserMessage(text=message_text))
        reply = response if isinstance(response, str) else str(response)
    except Exception as e:
        logger.error(f"WhatsApp GPT error: {e}")
        reply = "¡Hola! Bienvenido a Fakulti Laboratorios. Soy tu asesor virtual. ¿En qué puedo ayudarte?"
    
    # Parse lead name
    name_match = re.search(r'\[LEAD_NAME:([^\]]+)\]', reply)
    if name_match and not lead_name:
        detected_name = name_match.group(1).strip()
        await db.leads.update_one({"id": lead_id}, {"$set": {"name": detected_name, "last_interaction": datetime.now(timezone.utc).isoformat()}})
        await db.chat_sessions_meta.update_one({"session_id": session_id}, {"$set": {"lead_name": detected_name}}, upsert=True)
        logger.info(f"WhatsApp lead name detected: {detected_name} for {phone}")
    reply = re.sub(r'\[LEAD_NAME:[^\]]+\]', '', reply)
    
    # Parse lead data updates
    update_matches = re.findall(r'\[UPDATE_LEAD:(\w+)=([^\]]+)\]', reply)
    if update_matches:
        update_fields = {}
        allowed_fields = {"city", "product_interest", "email", "ci_ruc", "address"}
        for field, value in update_matches:
            if field in allowed_fields:
                update_fields[field] = value.strip()
        if update_fields:
            update_fields["last_interaction"] = datetime.now(timezone.utc).isoformat()
            await db.leads.update_one({"id": lead_id}, {"$set": update_fields})
            logger.info(f"WhatsApp lead {lead_id} updated: {update_fields}")
    reply = re.sub(r'\[UPDATE_LEAD:\w+=[^\]]+\]', '', reply)
    
    # Parse stage classification
    stage_match = re.search(r'\[STAGE:(\w+)\]', reply)
    if stage_match:
        new_stage = stage_match.group(1).strip()
        if new_stage in FUNNEL_STAGES:
            current_stage = existing_lead.get("funnel_stage", "nuevo")
            current_priority = FUNNEL_STAGES.index(current_stage) if current_stage in FUNNEL_STAGES else 0
            new_priority = FUNNEL_STAGES.index(new_stage)
            if new_stage == "perdido" or new_priority > current_priority:
                update_data = {"funnel_stage": new_stage, "last_interaction": datetime.now(timezone.utc).isoformat()}
                # Hot-lead detection: mark as needs_advisor when reaching buying stages
                if new_stage in ("en_negociacion", "cliente_nuevo") and not existing_lead.get("assigned_advisor"):
                    update_data["needs_advisor"] = True
                    # Create admin notification for hot lead
                    existing_hot_notif = await db.advisor_notifications.find_one(
                        {"lead_id": lead_id, "type": "hot_lead", "read": False}, {"_id": 0}
                    )
                    if not existing_hot_notif:
                        lead_name_for_notif = existing_lead.get("name", phone)
                        product_for_notif = existing_lead.get("product_interest", "")
                        stage_label = "quiere comprar" if new_stage == "cliente_nuevo" else "en negociación"
                        await db.advisor_notifications.insert_one({
                            "id": str(uuid.uuid4()),
                            "advisor_id": "admin",
                            "type": "hot_lead",
                            "title": f"Lead caliente: {lead_name_for_notif} ({stage_label})",
                            "message": f"{lead_name_for_notif} está listo para compra{f' de {product_for_notif}' if product_for_notif else ''}. Asigna un asesor para cerrar la venta.",
                            "lead_id": lead_id,
                            "session_id": session_id,
                            "read": False,
                            "created_at": datetime.now(timezone.utc).isoformat()
                        })
                        logger.info(f"Hot lead notification: {lead_name_for_notif} reached {new_stage}")
                await db.leads.update_one({"id": lead_id}, {"$set": update_data})
    reply = re.sub(r'\[STAGE:\w+\]', '', reply).strip()
    
    # ===== ADVISOR HANDOVER DETECTION =====
    # Detect if USER asked for an advisor
    msg_lower = message_text.strip().lower()
    user_wants_handover = any(kw in msg_lower for kw in HANDOVER_KEYWORDS)
    
    # Detect if BOT decided to transfer to an advisor
    BOT_TRANSFER_PHRASES = [
        "transfiero con un asesor", "te transfiero", "comunico con un asesor",
        "asesor se comunicará", "asesor se comunicara", "asesor te contactará",
        "asesor te contactara", "paso con un asesor", "derivo con un asesor",
        "un asesor te atenderá", "un asesor te atendera", "contactar con un asesor",
        "te paso con un humano", "te comunico con alguien",
        "aviso a un asesor", "asesor te escriba", "asesor del equipo",
        "te contactará un asesor", "te contactara un asesor",
        "asesor se pondrá en contacto", "asesor se pondra en contacto",
        "conecto con un asesor", "te conecto con", "conectar con un asesor",
        "asesor humano"
    ]
    reply_lower = reply.lower()
    bot_wants_handover = any(phrase in reply_lower for phrase in BOT_TRANSFER_PHRASES)
    
    needs_handover = user_wants_handover or bot_wants_handover
    handover_reason = "solicitud_usuario" if user_wants_handover else ("bot_transfer" if bot_wants_handover else None)
    
    if needs_handover and lead_id:
        # Mark lead as needing advisor
        await db.leads.update_one(
            {"id": lead_id},
            {"$set": {"needs_advisor": True, "last_interaction": datetime.now(timezone.utc).isoformat()}}
        )
        # Create handover alert if none exists
        existing_alert = await db.handover_alerts.find_one({"lead_id": lead_id, "status": "pending"}, {"_id": 0})
        if not existing_alert:
            lead_doc = await db.leads.find_one({"id": lead_id}, {"_id": 0, "name": 1, "whatsapp": 1, "product_interest": 1, "channel": 1, "funnel_stage": 1})
            lead_name_alert = (lead_doc.get("name", "") if lead_doc else "") or phone
            product_alert = (lead_doc.get("product_interest", "") if lead_doc else "")
            alert_message = message_text[:150] if user_wants_handover else f"Bot transfirió: {reply[:120]}"
            await db.handover_alerts.insert_one({
                "id": str(uuid.uuid4()),
                "lead_id": lead_id,
                "lead_name": lead_name_alert,
                "lead_phone": phone,
                "message": alert_message,
                "reason": handover_reason,
                "product": product_alert,
                "channel": (lead_doc.get("channel", "") if lead_doc else ""),
                "status": "pending",
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            logger.info(f"Handover alert created for {lead_name_alert} ({phone}) - reason: {handover_reason}")
        # Create admin notification for immediate visibility
        existing_notif = await db.advisor_notifications.find_one(
            {"lead_id": lead_id, "type": "advisor_request", "read": False}, {"_id": 0}
        )
        if not existing_notif:
            lead_doc_n = await db.leads.find_one({"id": lead_id}, {"_id": 0, "name": 1, "product_interest": 1})
            lead_name_notif = (lead_doc_n.get("name", "") if lead_doc_n else "") or phone
            product_notif = (lead_doc_n.get("product_interest", "") if lead_doc_n else "")
            reason_label = "El cliente solicitó un asesor" if user_wants_handover else "El bot transfirió al cliente"
            await db.advisor_notifications.insert_one({
                "id": str(uuid.uuid4()),
                "advisor_id": "admin",
                "type": "advisor_request",
                "title": f"Solicitud de asesor: {lead_name_notif}",
                "message": f"{reason_label}{f' para {product_notif}' if product_notif else ''}. Requiere atención inmediata.",
                "lead_id": lead_id,
                "session_id": session_id,
                "read": False,
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            logger.info(f"Advisor request notification for {lead_name_notif} (reason: {handover_reason})")
    
    # Update last interaction
    await db.leads.update_one({"id": lead_id}, {"$set": {"last_interaction": datetime.now(timezone.utc).isoformat()}})
    
    return reply, lead_id

# Meta WhatsApp webhook verification (GET)
from fastapi import Request
from fastapi.responses import PlainTextResponse

@api_router.get("/webhook/whatsapp")
async def whatsapp_verify(request: Request):
    """Verification endpoint for Meta WhatsApp Cloud API webhook setup."""
    params = request.query_params
    mode = params.get("hub.mode")
    token = params.get("hub.verify_token")
    challenge = params.get("hub.challenge")
    config = await get_whatsapp_config()
    valid_token = config.get("verify_token", "") or os.environ.get("VERIFY_TOKEN", "")
    if mode == "subscribe" and token and token == valid_token:
        logger.info("WhatsApp webhook verified successfully")
        return PlainTextResponse(content=challenge, status_code=200)
    raise HTTPException(status_code=403, detail="Verification failed")

# Meta WhatsApp incoming message (POST)
@api_router.post("/webhook/whatsapp")
async def whatsapp_incoming(request: Request):
    """Receive incoming WhatsApp messages from Meta Cloud API."""
    try:
        body = await request.json()
    except Exception:
        return {"status": "ok"}
    
    entries = body.get("entry", [])
    for entry in entries:
        changes = entry.get("changes", [])
        for change in changes:
            value = change.get("value", {})
            # Ignore status updates (delivery receipts, read receipts)
            if value.get("statuses"):
                continue
            messages = value.get("messages", [])
            for msg in messages:
                phone = normalize_phone_ec(msg.get("from", ""))
                msg_type = msg.get("type", "")
                
                # Extract text content based on message type
                if msg_type == "text":
                    text = msg.get("text", {}).get("body", "")
                elif msg_type == "image":
                    text = msg.get("image", {}).get("caption", "") or "[El cliente envió una imagen]"
                elif msg_type == "audio":
                    text = "[El cliente envió un audio]"
                elif msg_type == "video":
                    text = msg.get("video", {}).get("caption", "") or "[El cliente envió un video]"
                elif msg_type == "document":
                    text = f"[El cliente envió un documento: {msg.get('document', {}).get('filename', 'archivo')}]"
                elif msg_type == "sticker":
                    text = "[El cliente envió un sticker]"
                elif msg_type == "location":
                    loc = msg.get("location", {})
                    text = f"[Ubicación: {loc.get('latitude', '')}, {loc.get('longitude', '')}]"
                elif msg_type == "contacts":
                    text = "[El cliente compartió un contacto]"
                elif msg_type == "reaction":
                    continue  # Skip reactions silently
                else:
                    text = f"[Mensaje tipo: {msg_type}]"
                
                if phone and text:
                        logger.info(f"WhatsApp incoming from {phone}: {text[:50]}...")
                        start_time = datetime.now(timezone.utc)
                        reply, lead_id = await process_whatsapp_incoming(phone, text)
                        response_time_ms = int((datetime.now(timezone.utc) - start_time).total_seconds() * 1000)
                        sent = await send_whatsapp_message(phone, reply)
                        # Store in chat history
                        session_id = f"wa_{phone}"
                        now = datetime.now(timezone.utc).isoformat()
                        # Get the updated lead name (may have been detected in this interaction)
                        updated_lead = await db.leads.find_one({"id": lead_id}, {"_id": 0, "name": 1}) if lead_id else None
                        resolved_lead_name = (updated_lead.get("name", "") if updated_lead else "") or phone
                        await db.chat_messages.insert_one({"id": str(uuid.uuid4()), "session_id": session_id, "lead_id": lead_id, "role": "user", "content": text, "timestamp": now, "source": "whatsapp"})
                        await db.chat_messages.insert_one({"id": str(uuid.uuid4()), "session_id": session_id, "lead_id": lead_id, "role": "assistant", "content": reply, "timestamp": now, "source": "whatsapp", "response_time_ms": response_time_ms, "delivered": sent})
                        await db.chat_sessions_meta.update_one(
                            {"session_id": session_id},
                            {"$set": {"session_id": session_id, "lead_id": lead_id, "lead_name": resolved_lead_name, "source": "whatsapp", "last_activity": now}},
                            upsert=True
                        )
                        # Detect bot timeout: if lead has been interacting for over 1 min without stage progress
                        if lead_id:
                            session_id_check = f"wa_{phone}"
                            recent_msgs = await db.chat_messages.find(
                                {"session_id": session_id_check, "role": "user"},
                                {"_id": 0, "timestamp": 1}
                            ).sort("timestamp", -1).to_list(5)
                            if len(recent_msgs) >= 3:
                                first_msg_time = recent_msgs[-1].get("timestamp", "")
                                if first_msg_time:
                                    try:
                                        first_dt = datetime.fromisoformat(first_msg_time.replace("Z", "+00:00"))
                                        now_dt = datetime.now(timezone.utc)
                                        elapsed = (now_dt - first_dt).total_seconds()
                                        if elapsed > BOT_TIMEOUT_SECONDS:
                                            lead_check = await db.leads.find_one({"id": lead_id}, {"_id": 0, "funnel_stage": 1})
                                            if lead_check and lead_check.get("funnel_stage") == "nuevo":
                                                existing_timeout_alert = await db.handover_alerts.find_one({"lead_id": lead_id, "status": "pending"}, {"_id": 0})
                                                if not existing_timeout_alert:
                                                    lead_doc_t = await db.leads.find_one({"id": lead_id}, {"_id": 0, "name": 1, "whatsapp": 1, "product_interest": 1, "channel": 1})
                                                    await db.handover_alerts.insert_one({
                                                        "id": str(uuid.uuid4()),
                                                        "lead_id": lead_id,
                                                        "lead_name": (lead_doc_t.get("name", "") if lead_doc_t else "") or phone,
                                                        "lead_phone": phone,
                                                        "message": text,
                                                        "reason": "timeout_bot",
                                                        "product": (lead_doc_t.get("product_interest", "") if lead_doc_t else ""),
                                                        "channel": (lead_doc_t.get("channel", "") if lead_doc_t else ""),
                                                        "status": "pending",
                                                        "created_at": datetime.now(timezone.utc).isoformat()
                                                    })
                                                    await db.leads.update_one({"id": lead_id}, {"$set": {"needs_advisor": True}})
                                                    logger.info(f"Timeout handover alert for {phone}")
                                    except Exception:
                                        pass
                        
                        # Notify assigned advisor when their lead writes
                        if lead_id:
                            lead_check_advisor = await db.leads.find_one({"id": lead_id}, {"_id": 0, "assigned_advisor": 1, "name": 1})
                            if lead_check_advisor and lead_check_advisor.get("assigned_advisor"):
                                existing_notif = await db.advisor_notifications.find_one(
                                    {"lead_id": lead_id, "advisor_id": lead_check_advisor["assigned_advisor"], "read": False}, {"_id": 0}
                                )
                                if not existing_notif:
                                    await db.advisor_notifications.insert_one({
                                        "id": str(uuid.uuid4()),
                                        "advisor_id": lead_check_advisor["assigned_advisor"],
                                        "lead_id": lead_id,
                                        "lead_name": lead_check_advisor.get("name", ""),
                                        "lead_phone": phone,
                                        "message": text[:100],
                                        "type": "new_message",
                                        "read": False,
                                        "created_at": datetime.now(timezone.utc).isoformat()
                                    })
                                    logger.info(f"Notification created for advisor {lead_check_advisor['assigned_advisor']} - lead {lead_id} wrote")
    return {"status": "ok"}

# Legacy internal webhook (for Chat IA testing)
class WhatsAppMessage(BaseModel):
    from_number: str
    message: str

@api_router.post("/whatsapp/webhook")
async def whatsapp_webhook_legacy(req: WhatsAppMessage):
    phone = normalize_phone_ec(req.from_number.strip())
    session_id = f"wa_{phone}"
    now = datetime.now(timezone.utc).isoformat()
    
    # Store user message before processing (so AI sees the full history)
    await db.chat_messages.insert_one({"id": str(uuid.uuid4()), "session_id": session_id, "lead_id": None, "role": "user", "content": req.message, "timestamp": now, "source": "whatsapp"})
    
    reply, lead_id = await process_whatsapp_incoming(phone, req.message)
    
    # Store bot reply
    await db.chat_messages.insert_one({"id": str(uuid.uuid4()), "session_id": session_id, "lead_id": lead_id, "role": "assistant", "content": reply, "timestamp": now, "source": "whatsapp"})
    
    # Update session meta
    updated_lead = await db.leads.find_one({"id": lead_id}, {"_id": 0, "name": 1}) if lead_id else None
    resolved_name = (updated_lead.get("name", "") if updated_lead else "") or phone
    await db.chat_sessions_meta.update_one(
        {"session_id": session_id},
        {"$set": {"session_id": session_id, "lead_id": lead_id, "lead_name": resolved_name, "source": "whatsapp", "last_activity": now}},
        upsert=True
    )
    
    # Update lead_id on the user message too
    if lead_id:
        await db.chat_messages.update_one({"session_id": session_id, "role": "user", "lead_id": None}, {"$set": {"lead_id": lead_id}})
    
    return {"reply": reply, "lead_id": lead_id}

# ========== AUTOMATION RULES ==========

@api_router.get("/automation/rules")
async def get_automation_rules(user=Depends(get_current_user)):
    rules = await db.automation_rules.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    return rules

@api_router.post("/automation/rules")
async def create_automation_rule(req: AutomationRuleCreate, user=Depends(get_current_user)):
    count = await db.automation_rules.count_documents({})
    doc = {"id": str(uuid.uuid4()), **req.model_dump(), "order": count + 1, "created_at": datetime.now(timezone.utc).isoformat()}
    await db.automation_rules.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/automation/rules/{rule_id}")
async def update_automation_rule(rule_id: str, req: AutomationRuleCreate, user=Depends(get_current_user)):
    await db.automation_rules.update_one({"id": rule_id}, {"$set": req.model_dump()})
    rule = await db.automation_rules.find_one({"id": rule_id}, {"_id": 0})
    return rule

@api_router.patch("/automation/rules/{rule_id}/toggle")
async def toggle_automation_rule(rule_id: str, user=Depends(get_current_user)):
    rule = await db.automation_rules.find_one({"id": rule_id}, {"_id": 0})
    if not rule:
        raise HTTPException(status_code=404, detail="Regla no encontrada")
    new_active = not rule.get("active", True)
    await db.automation_rules.update_one({"id": rule_id}, {"$set": {"active": new_active}})
    return {"active": new_active}

@api_router.delete("/automation/rules/{rule_id}")
async def delete_automation_rule(rule_id: str, user=Depends(get_current_user)):
    await db.automation_rules.delete_one({"id": rule_id})
    return {"message": "Regla eliminada"}

# ========== WHATSAPP CONFIG ==========

@api_router.get("/config/whatsapp")
async def get_wa_config(user=Depends(get_current_user)):
    config = await get_whatsapp_config()
    safe = {**config}
    if safe.get("access_token"):
        safe["access_token"] = safe["access_token"][:10] + "..." + safe["access_token"][-4:] if len(safe["access_token"]) > 14 else "****"
    return safe

@api_router.put("/config/whatsapp")
async def update_wa_config(req: WhatsAppConfigUpdate, user=Depends(get_current_user)):
    data = req.model_dump()
    # Don't overwrite token if masked value sent
    if data.get("access_token") and "..." in data["access_token"]:
        existing = await get_whatsapp_config()
        data["access_token"] = existing.get("access_token", "")
    data["id"] = "main"
    await db.whatsapp_config.update_one({"id": "main"}, {"$set": data}, upsert=True)
    config = await db.whatsapp_config.find_one({"id": "main"}, {"_id": 0})
    if config.get("access_token"):
        config["access_token"] = config["access_token"][:10] + "..." + config["access_token"][-4:] if len(config["access_token"]) > 14 else "****"
    return config

@api_router.post("/config/whatsapp/test")
async def test_wa_connection(user=Depends(get_current_user)):
    config = await get_whatsapp_config()
    if not config.get("phone_number_id") or not config.get("access_token"):
        return {"success": False, "message": "Credenciales no configuradas"}
    url = f"{WHATSAPP_API_URL}/{config['phone_number_id']}"
    headers = {"Authorization": f"Bearer {config['access_token']}"}
    try:
        async with httpx.AsyncClient() as c:
            resp = await c.get(url, headers=headers, timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                return {"success": True, "message": f"Conectado: {data.get('display_phone_number', 'OK')}", "phone": data.get("display_phone_number")}
            return {"success": False, "message": f"Error {resp.status_code}: {resp.text[:200]}"}
    except Exception as e:
        return {"success": False, "message": str(e)}

# ========== AI CONFIG ==========

@api_router.get("/config/ai")
async def get_ai_config(user=Depends(get_current_user)):
    config = await db.ai_config.find_one({"id": "main"}, {"_id": 0})
    return config or {"id": "main", "intent_analysis": True, "lead_classification": True, "product_recommendation": True, "suggested_responses": True}

@api_router.put("/config/ai")
async def update_ai_config(req: AIConfigUpdate, user=Depends(get_current_user)):
    data = req.model_dump()
    data["id"] = "main"
    await db.ai_config.update_one({"id": "main"}, {"$set": data}, upsert=True)
    return await db.ai_config.find_one({"id": "main"}, {"_id": 0})

# ========== HUMAN AGENT DERIVATION ==========

# ========== QR CAMPAIGNS ==========

@api_router.get("/qr-campaigns")
async def get_qr_campaigns(user=Depends(get_current_user)):
    campaigns = await db.qr_campaigns.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    for c in campaigns:
        c["leads_count"] = await db.leads.count_documents({"channel": c.get("channel", ""), "source": c.get("source", "")})
        c.setdefault("scan_count", 0)
    return campaigns

@api_router.post("/qr-campaigns")
async def create_qr_campaign(req: QRCampaignCreate, user=Depends(get_current_user)):
    doc = {
        "id": str(uuid.uuid4()),
        **req.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.qr_campaigns.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/qr-campaigns/{campaign_id}")
async def update_qr_campaign(campaign_id: str, req: QRCampaignCreate, user=Depends(get_current_user)):
    await db.qr_campaigns.update_one({"id": campaign_id}, {"$set": req.model_dump()})
    campaign = await db.qr_campaigns.find_one({"id": campaign_id}, {"_id": 0})
    return campaign

@api_router.delete("/qr-campaigns/{campaign_id}")
async def delete_qr_campaign(campaign_id: str, user=Depends(get_current_user)):
    await db.qr_campaigns.delete_one({"id": campaign_id})
    return {"message": "Campaña QR eliminada"}

@api_router.get("/qr-campaigns/{campaign_id}/qrcode")
async def generate_qr_code(campaign_id: str):
    """Generate a QR code image for a campaign's WhatsApp link."""
    import qrcode
    from PIL import Image
    
    campaign = await db.qr_campaigns.find_one({"id": campaign_id}, {"_id": 0})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaña no encontrada")
    
    # Get WhatsApp phone number from config
    config = await get_whatsapp_config()
    wa_phone = config.get("display_phone", "")
    if not wa_phone:
        # Try getting from Meta API
        if config.get("phone_number_id") and config.get("access_token"):
            try:
                url = f"{WHATSAPP_API_URL}/{config['phone_number_id']}"
                headers = {"Authorization": f"Bearer {config['access_token']}"}
                async with httpx.AsyncClient() as c:
                    resp = await c.get(url, headers=headers, timeout=10)
                    if resp.status_code == 200:
                        wa_phone = resp.json().get("display_phone_number", "")
            except Exception:
                pass
        if not wa_phone:
            wa_phone = "593000000000"
    
    # Clean phone for wa.me link
    wa_phone_clean = wa_phone.replace("+", "").replace(" ", "").replace("-", "")
    if wa_phone_clean.startswith("0"):
        wa_phone_clean = "593" + wa_phone_clean[1:]
    
    # Build WhatsApp link with pre-filled message
    import urllib.parse
    encoded_msg = urllib.parse.quote(campaign["initial_message"])
    wa_link = f"https://wa.me/{wa_phone_clean}?text={encoded_msg}"
    
    # Generate QR code
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=10, border=4)
    qr.add_data(wa_link)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#1A6B3C", back_color="white").convert("RGB")
    
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="image/png",
        headers={"Content-Disposition": f"attachment; filename=qr_{campaign['name'].replace(' ', '_')}.png"}
    )

@api_router.get("/qr-campaigns/{campaign_id}/link")
async def get_qr_link(campaign_id: str, user=Depends(get_current_user)):
    """Get the WhatsApp link for a QR campaign."""
    campaign = await db.qr_campaigns.find_one({"id": campaign_id}, {"_id": 0})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaña no encontrada")
    
    config = await get_whatsapp_config()
    wa_phone = config.get("display_phone", "593000000000")
    wa_phone_clean = wa_phone.replace("+", "").replace(" ", "").replace("-", "")
    if wa_phone_clean.startswith("0"):
        wa_phone_clean = "593" + wa_phone_clean[1:]
    
    import urllib.parse
    encoded_msg = urllib.parse.quote(campaign["initial_message"])
    wa_link = f"https://wa.me/{wa_phone_clean}?text={encoded_msg}"
    
    return {"link": wa_link, "campaign": campaign}

# ========== INITIAL INTENTS ==========


# ========== QR/CHANNEL AUTO-DETECTION HELPER ==========

async def detect_channel_from_message(message_text: str, lead_id: str):
    """Check if the incoming message matches a QR campaign and auto-tag the lead."""
    msg_lower = message_text.strip().lower()
    
    # Check QR campaigns (exact match on initial message)
    campaigns = await db.qr_campaigns.find({"active": True}, {"_id": 0}).to_list(50)
    for campaign in campaigns:
        campaign_msg = campaign["initial_message"].strip().lower()
        if msg_lower == campaign_msg or campaign_msg in msg_lower:
            update = {
                "channel": campaign.get("channel", ""),
                "source": campaign.get("source", ""),
                "last_interaction": datetime.now(timezone.utc).isoformat()
            }
            if campaign.get("product"):
                update["product_interest"] = campaign["product"]
            await db.leads.update_one({"id": lead_id}, {"$set": update})
            # Increment scan count
            await db.qr_campaigns.update_one({"id": campaign["id"]}, {"$inc": {"scan_count": 1}})
            logger.info(f"QR campaign matched for lead {lead_id}: {campaign['name']} -> channel={campaign['channel']} (scan #{campaign.get('scan_count', 0) + 1})")
            return True
    
    return False

@api_router.post("/leads/{lead_id}/derive-human")
async def derive_to_human(lead_id: str, reason: str = Query("general"), user=Depends(get_current_user)):
    await db.leads.update_one(
        {"id": lead_id},
        {"$set": {"status": "gestion_humana", "notes": f"Derivado a agente humano: {reason}", "last_interaction": datetime.now(timezone.utc).isoformat()}}
    )
    notification = {
        "id": str(uuid.uuid4()),
        "type": "derivacion_humana",
        "lead_id": lead_id,
        "reason": reason,
        "read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one(notification)
    return {"message": "Lead derivado a agente humano"}

@api_router.get("/notifications")
async def get_notifications(user=Depends(get_current_user)):
    notifs = await db.notifications.find({}, {"_id": 0}).sort("created_at", -1).limit(50).to_list(50)
    return notifs


# ========== BLOCK 9: AUTO-ENROLLMENT CONFIGURATION ==========

@api_router.post("/loyalty/auto-enroll-config")
async def set_auto_enroll_config(body: dict, user=Depends(get_current_user)):
    """Configure auto-enrollment: when lead reaches a target stage, auto-enroll in a sequence."""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")
    config = {
        "id": "auto_enroll_config",
        "enabled": body.get("enabled", False),
        "target_stage": body.get("target_stage", "cliente_nuevo"),
        "default_sequence_id": body.get("default_sequence_id", ""),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.system_config.update_one({"id": "auto_enroll_config"}, {"$set": config}, upsert=True)
    return config

@api_router.get("/loyalty/auto-enroll-config")
async def get_auto_enroll_config(user=Depends(get_current_user)):
    config = await db.system_config.find_one({"id": "auto_enroll_config"}, {"_id": 0})
    return config or {"id": "auto_enroll_config", "enabled": False, "target_stage": "cliente_nuevo", "default_sequence_id": ""}

# ========== BLOCK 10: PROMOTIONS & CAMPAIGNS ==========

class CampaignCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    campaign_type: Optional[str] = "promo"
    target_stage: Optional[str] = ""
    target_product: Optional[str] = ""
    target_channel: Optional[str] = ""
    target_season: Optional[str] = ""
    message_template: str
    image_url: Optional[str] = ""
    scheduled_date: Optional[str] = ""
    active: Optional[bool] = True

@api_router.get("/campaigns")
async def get_campaigns(user=Depends(get_current_user)):
    campaigns = await db.campaigns.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return campaigns

@api_router.post("/campaigns")
async def create_campaign(req: CampaignCreate, user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")
    # Count matching leads
    query = {}
    if req.target_stage:
        query["funnel_stage"] = req.target_stage
    if req.target_product:
        query["product_interest"] = {"$regex": req.target_product, "$options": "i"}
    if req.target_channel:
        query["channel"] = req.target_channel
    if req.target_season:
        query["season"] = req.target_season
    target_count = await db.leads.count_documents(query)
    
    campaign = {
        "id": str(uuid.uuid4()),
        **req.model_dump(),
        "target_count": target_count,
        "sent_count": 0,
        "failed_count": 0,
        "status": "draft",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.campaigns.insert_one(campaign)
    campaign.pop("_id", None)
    return campaign

@api_router.put("/campaigns/{campaign_id}")
async def update_campaign(campaign_id: str, req: CampaignCreate, user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")
    update = {k: v for k, v in req.model_dump().items() if v is not None}
    # Recalculate target_count based on new filters
    query = {}
    if req.target_stage:
        query["funnel_stage"] = req.target_stage
    if req.target_product:
        query["product_interest"] = {"$regex": req.target_product, "$options": "i"}
    if req.target_channel:
        query["channel"] = req.target_channel
    update["target_count"] = await db.leads.count_documents(query)
    await db.campaigns.update_one({"id": campaign_id}, {"$set": update})
    campaign = await db.campaigns.find_one({"id": campaign_id}, {"_id": 0})
    return campaign

@api_router.delete("/campaigns/{campaign_id}")
async def delete_campaign(campaign_id: str, user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")
    await db.campaigns.delete_one({"id": campaign_id})
    return {"message": "Campaña eliminada"}

@api_router.post("/campaigns/{campaign_id}/send")
async def send_campaign(campaign_id: str, body: dict = {}, user=Depends(get_current_user)):
    """Send campaign messages in batches. Records each message in chat history."""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")
    campaign = await db.campaigns.find_one({"id": campaign_id}, {"_id": 0})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaña no encontrada")
    
    batch_size = body.get("batch_size", 50)
    query = {}
    if campaign.get("target_stage"):
        query["funnel_stage"] = campaign["target_stage"]
    if campaign.get("target_product"):
        query["product_interest"] = {"$regex": campaign["target_product"], "$options": "i"}
    if campaign.get("target_channel"):
        query["channel"] = campaign["target_channel"]
    
    leads = await db.leads.find(query, {"_id": 0, "id": 1, "name": 1, "whatsapp": 1}).to_list(500)
    
    # Update target_count with real-time data
    await db.campaigns.update_one({"id": campaign_id}, {"$set": {"target_count": len(leads)}})
    
    wa_config = await get_whatsapp_config()
    sent = 0
    failed = 0
    now_iso = datetime.now(timezone.utc).isoformat()
    
    for lead in leads[:batch_size]:
        try:
            msg = campaign["message_template"].replace("{nombre}", lead.get("name", "").split()[0] if lead.get("name") else "")
            lead_id = lead["id"]
            image_url = campaign.get("image_url", "")
            
            # 1. Find or create chat session for this lead (prefer wa_ session)
            phone = lead.get("whatsapp", "")
            wa_session_id = f"wa_{phone}" if phone else None
            meta = None
            if wa_session_id:
                meta = await db.chat_sessions_meta.find_one({"session_id": wa_session_id}, {"_id": 0})
            if not meta:
                meta = await db.chat_sessions_meta.find_one({"lead_id": lead_id, "source": "whatsapp"}, {"_id": 0})
            if not meta:
                meta = await db.chat_sessions_meta.find_one({"lead_id": lead_id}, {"_id": 0})
            if not meta:
                session_id = wa_session_id or f"lead_{lead_id}_{int(datetime.now(timezone.utc).timestamp())}"
                await db.chat_sessions_meta.insert_one({
                    "session_id": session_id, "lead_id": lead_id,
                    "lead_name": lead.get("name", ""), "lead_phone": phone, "source": "whatsapp"
                })
            else:
                session_id = meta["session_id"]
            
            # 2. Record the campaign message in chat history (only message + image)
            content = msg
            if image_url:
                content += f"\n[Imagen: {image_url}]"
            chat_msg = {
                "id": str(uuid.uuid4()),
                "session_id": session_id,
                "role": "assistant",
                "content": content,
                "timestamp": now_iso,
                "source": "campaign"
            }
            await db.chat_messages.insert_one(chat_msg)
            
            # 3. Send via WhatsApp API if configured
            if wa_config and wa_config.get("phone_number_id") and wa_config.get("access_token"):
                wa_phone = phone
                
                if image_url:
                    # Send image with caption via WhatsApp media message
                    full_image_url = image_url
                    if image_url.startswith("/api/") or image_url.startswith("/"):
                        public_url = os.environ.get("PUBLIC_URL", "").rstrip("/")
                        full_image_url = f"{public_url}{image_url}"
                    logger.info(f"Campaign image URL resolved: {full_image_url}")
                    await send_whatsapp_image(wa_phone, full_image_url, msg)
                else:
                    # Send text only
                    await send_whatsapp_message(wa_phone, msg)
            
            # 4. Update lead last_interaction
            await db.leads.update_one({"id": lead_id}, {"$set": {"last_interaction": now_iso}})
            
            sent += 1
        except Exception as e:
            logger.error(f"Campaign send error for lead {lead.get('id')}: {e}")
            failed += 1
    
    await db.campaigns.update_one({"id": campaign_id}, {
        "$set": {"status": "sent", "sent_count": campaign.get("sent_count", 0) + sent, "failed_count": campaign.get("failed_count", 0) + failed,
                 "last_sent_at": now_iso}
    })
    return {"message": f"Campaña enviada: {sent} exitosos, {failed} fallidos de {len(leads[:batch_size])} leads", "sent": sent, "failed": failed}

# ========== BLOCK 11: SMART REMINDERS ==========

REMINDER_TEMPLATES_BY_STAGE = {
    "cliente_nuevo": [
        "Hola {nombre} 😊 Esperamos que estés disfrutando tu producto. Si necesitas algo adicional o tienes alguna consulta, estamos para ayudarte.",
        "Hola {nombre}, queremos saber cómo te fue con tu pedido. ¿Necesitas algo más? Estamos a tu disposición.",
    ],
    "cliente_activo": [
        "Hola {nombre} 😊 ¿Cómo te ha ido con tu producto? Si necesitas asesoría o reposición, aquí estamos para ti.",
        "Hola {nombre}, esperamos que todo vaya bien. Si necesitas algo más, no dudes en escribirnos.",
    ],
    "en_negociacion": [
        "Hola {nombre} 😊 Vimos que estabas interesado en {producto}. ¿Tienes alguna duda que podamos resolver para avanzar con tu pedido?",
        "Hola {nombre}, solo queríamos saber si pudiste decidirte por {producto}. Estamos listos para ayudarte a concretar tu compra.",
    ],
    "interesado": [
        "Hola {nombre} 😊 Notamos que te interesaba {producto}. ¿Te gustaría más información o tienes alguna pregunta?",
        "Hola {nombre}, seguimos con disponibilidad de {producto}. ¿Quieres que te cuente más sobre sus beneficios?",
    ],
    "nuevo": [
        "Hola {nombre} 😊 Soy el asesor virtual de Fakulti. ¿En qué puedo ayudarte hoy?",
        "Hola {nombre}, ¿hay algún producto de Fakulti que te interese? Con gusto te asesoro.",
    ],
}

class ReminderCreate(BaseModel):
    name: str
    message_template: Optional[str] = ""
    target_stage: Optional[str] = ""
    target_product: Optional[str] = ""
    days_since_last_interaction: Optional[int] = 7
    batch_size: Optional[int] = 10
    active: Optional[bool] = True

@api_router.get("/reminders")
async def get_reminders(user=Depends(get_current_user)):
    reminders = await db.reminders.find({}, {"_id": 0}).sort("created_at", -1).to_list(50)
    return reminders

@api_router.post("/reminders")
async def create_reminder(req: ReminderCreate, user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")
    reminder = {
        "id": str(uuid.uuid4()),
        **req.model_dump(),
        "last_run": None,
        "total_sent": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.reminders.insert_one(reminder)
    reminder.pop("_id", None)
    return reminder

@api_router.delete("/reminders/{reminder_id}")
async def delete_reminder(reminder_id: str, user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")
    await db.reminders.delete_one({"id": reminder_id})
    return {"message": "Recordatorio eliminado"}

def _build_smart_reminder_message(lead: dict, custom_template: str = "") -> str:
    """Build a context-aware reminder message based on lead stage."""
    name = lead.get("name", "").split()[0] if lead.get("name") else ""
    stage = lead.get("funnel_stage", "nuevo")
    product = lead.get("product_interest", "nuestros productos")

    if custom_template:
        return custom_template.replace("{nombre}", name).replace("{producto}", product)

    templates = REMINDER_TEMPLATES_BY_STAGE.get(stage, REMINDER_TEMPLATES_BY_STAGE["nuevo"])
    import random as _rnd
    template = _rnd.choice(templates)
    return template.replace("{nombre}", name).replace("{producto}", product)

@api_router.post("/reminders/{reminder_id}/execute")
async def execute_reminder(reminder_id: str, user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")
    reminder = await db.reminders.find_one({"id": reminder_id}, {"_id": 0})
    if not reminder:
        raise HTTPException(status_code=404, detail="Recordatorio no encontrado")
    
    cutoff = (datetime.now(timezone.utc) - timedelta(days=reminder.get("days_since_last_interaction", 7))).isoformat()
    query = {"last_interaction": {"$lte": cutoff}, "whatsapp": {"$ne": ""}}
    if reminder.get("target_stage"):
        query["funnel_stage"] = reminder["target_stage"]
    if reminder.get("target_product"):
        query["product_interest"] = {"$regex": reminder["target_product"], "$options": "i"}
    
    batch = reminder.get("batch_size", 10)
    leads = await db.leads.find(query, {"_id": 0, "id": 1, "name": 1, "whatsapp": 1, "funnel_stage": 1, "product_interest": 1}).limit(batch).to_list(batch)
    
    wa_config = await get_whatsapp_config()
    sent = 0
    for lead in leads:
        try:
            msg = _build_smart_reminder_message(lead, reminder.get("message_template", ""))
            if wa_config and wa_config.get("phone_number_id") and wa_config.get("access_token"):
                await send_whatsapp_message(lead.get("whatsapp", ""), msg)
            
            # Store reminder in chat history
            session_id = f"wa_{lead.get('whatsapp', '')}"
            now = datetime.now(timezone.utc).isoformat()
            await db.chat_messages.insert_one({
                "id": str(uuid.uuid4()), "session_id": session_id, "lead_id": lead.get("id"),
                "role": "assistant", "content": msg, "timestamp": now,
                "source": "reminder"
            })
            await db.chat_sessions_meta.update_one(
                {"session_id": session_id},
                {"$set": {"last_activity": now}},
                upsert=True
            )
            await db.leads.update_one({"id": lead["id"]}, {"$set": {"last_interaction": now}})
            sent += 1
        except Exception:
            pass
    
    await db.reminders.update_one({"id": reminder_id}, {
        "$set": {"last_run": datetime.now(timezone.utc).isoformat(), "total_sent": reminder.get("total_sent", 0) + sent}
    })
    return {"message": f"Recordatorio ejecutado: {sent} mensajes enviados de {len(leads)} leads elegibles", "sent": sent}

# ========== BLOCK 12: ADMIN DASHBOARD BY ADVISOR ==========

@api_router.get("/dashboard/advisor-stats")
async def get_advisor_dashboard_stats(user=Depends(get_current_user)):
    """Get performance metrics per advisor for admin dashboard."""
    advisors = await db.admin_users.find({"role": "advisor"}, {"_id": 0, "password_hash": 0}).to_list(100)
    advisor_stats = []
    for a in advisors:
        aid = a["id"]
        total_leads = await db.leads.count_documents({"assigned_advisor": aid})
        won_leads = await db.leads.count_documents({"assigned_advisor": aid, "funnel_stage": {"$in": ["cliente_nuevo", "cliente_activo"]}})
        lost_leads = await db.leads.count_documents({"assigned_advisor": aid, "funnel_stage": "perdido"})
        negotiating = await db.leads.count_documents({"assigned_advisor": aid, "funnel_stage": "en_negociacion"})
        
        # Revenue from assigned leads
        pipeline = [
            {"$match": {"assigned_advisor": aid}},
            {"$unwind": "$purchase_history"},
            {"$group": {"_id": None, "total": {"$sum": "$purchase_history.price"}, "count": {"$sum": 1}}}
        ]
        rev = await db.leads.aggregate(pipeline).to_list(1)
        revenue = round(rev[0]["total"], 2) if rev else 0
        orders = rev[0]["count"] if rev else 0
        
        # Active conversations (bot paused = human control)
        active_chats = await db.leads.count_documents({"assigned_advisor": aid, "bot_paused": True})
        
        conversion = round((won_leads / total_leads * 100) if total_leads > 0 else 0, 1)
        
        advisor_stats.append({
            "id": aid,
            "name": a.get("name", ""),
            "email": a.get("email", ""),
            "status": a.get("status", "desconectado"),
            "specialization": a.get("specialization", ""),
            "total_leads": total_leads,
            "won_leads": won_leads,
            "lost_leads": lost_leads,
            "negotiating": negotiating,
            "revenue": revenue,
            "orders": orders,
            "active_chats": active_chats,
            "conversion_rate": conversion
        })
    
    # Global totals
    total_assigned = sum(a["total_leads"] for a in advisor_stats)
    total_unassigned = await db.leads.count_documents({"$or": [{"assigned_advisor": ""}, {"assigned_advisor": {"$exists": False}}]})
    total_revenue = sum(a["revenue"] for a in advisor_stats)
    
    return {
        "advisors": advisor_stats,
        "summary": {
            "total_assigned": total_assigned,
            "total_unassigned": total_unassigned,
            "total_advisors": len(advisors),
            "total_revenue_by_advisors": round(total_revenue, 2)
        }
    }

# ========== BLOCK 13: AI CONVERSATION ANALYSIS ==========

@api_router.post("/chat/analyze/{session_id}")
async def analyze_conversation(session_id: str, user=Depends(get_current_user)):
    """AI-powered conversation analysis: summary + suggested replies."""
    msgs = await db.chat_messages.find({"session_id": session_id}, {"_id": 0}).sort("timestamp", 1).to_list(100)
    if not msgs:
        raise HTTPException(status_code=404, detail="Sin mensajes para analizar")
    
    conversation_text = "\n".join([f"{'Cliente' if m['role'] == 'user' else 'Bot/Agente'}: {m['content']}" for m in msgs[-30:]])
    
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        import uuid as uuid_module
        
        system_message = """Eres un asistente de análisis de conversaciones para Fakulti (productos naturales).
Analiza la conversación entre un cliente y el bot/agente. Responde SIEMPRE en español con el siguiente formato JSON:
{
  "resumen": "Resumen conciso de la conversación (máximo 3 oraciones)",
  "sentimiento": "positivo|neutral|negativo",
  "interes_producto": "producto mencionado o 'no identificado'",
  "etapa_sugerida": "nuevo|interesado|en_negociacion|cliente_nuevo|perdido",
  "respuestas_sugeridas": ["respuesta sugerida 1", "respuesta sugerida 2", "respuesta sugerida 3"],
  "temas_clave": ["tema1", "tema2"],
  "nivel_urgencia": "alto|medio|bajo"
}"""
        
        api_key = os.environ.get("EMERGENT_LLM_KEY", "")
        llm_session_id = str(uuid_module.uuid4())
        
        llm = LlmChat(api_key=api_key, session_id=llm_session_id, system_message=system_message)
        llm = llm.with_model(provider="openai", model="gpt-5.2")
        
        user_message_text = f"Analiza esta conversación:\n\n{conversation_text}"
        user_message = UserMessage(text=user_message_text)
        response_text = await llm.send_message(user_message)
        
        import json as json_module
        try:
            # Try to parse as JSON
            text = response_text
            start = text.find("{")
            end = text.rfind("}") + 1
            if start >= 0 and end > start:
                analysis = json_module.loads(text[start:end])
            else:
                analysis = {"resumen": text, "sentimiento": "neutral", "respuestas_sugeridas": [], "temas_clave": [], "nivel_urgencia": "medio"}
        except Exception:
            analysis = {"resumen": response_text, "sentimiento": "neutral", "respuestas_sugeridas": [], "temas_clave": [], "nivel_urgencia": "medio"}
        
        return analysis
    except Exception as e:
        logger.error(f"AI analysis error: {e}")
        raise HTTPException(status_code=500, detail=f"Error en análisis IA: {str(e)}")

# ========== INCLUDE ROUTER & MIDDLEWARE ==========

app.include_router(api_router)

@app.get("/api/health")
async def health_check():
    return {"status": "ok"}

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve React static files in production (Railway)
import pathlib as _pathlib
_static_dir = _pathlib.Path(__file__).parent / "static"
if _static_dir.is_dir() and (_static_dir / "index.html").is_file():
    from starlette.staticfiles import StaticFiles
    from starlette.responses import FileResponse

    _static_assets = _static_dir / "static"
    if _static_assets.is_dir():
        app.mount("/static", StaticFiles(directory=str(_static_assets)), name="spa-static")

    @app.get("/{full_path:path}")
    async def serve_spa(full_path: str):
        if full_path.startswith("api/"):
            raise HTTPException(status_code=404)
        file_path = _static_dir / full_path
        if file_path.is_file() and ".." not in full_path:
            return FileResponse(file_path)
        return FileResponse(_static_dir / "index.html")

# ========== STARTUP - SEED DATA ==========

@app.on_event("startup")
async def startup():
    # Seed developer account
    dev_exists = await db.admin_users.find_one({"role": "developer"})
    if not dev_exists:
        try:
            dev_doc = {
                "id": str(uuid.uuid4()),
                "email": "aicreator@emprenovus.com",
                "password_hash": safe_hash_password("Jlsb*1082"),
                "name": "Desarrollador",
                "role": "developer",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.admin_users.insert_one(dev_doc)
            logger.info("Developer user seeded: aicreator@emprenovus.com")
        except Exception as e:
            logger.error(f"Failed to seed developer: {e}")
    
    admin_count = await db.admin_users.count_documents({"role": "admin"})
    if admin_count == 0:
        try:
            admin_doc = {
                "id": str(uuid.uuid4()),
                "email": "admin@fakulti.com",
                "password_hash": safe_hash_password("admin123"),
                "name": "Admin Fakulti",
                "role": "admin",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.admin_users.insert_one(admin_doc)
            logger.info("Admin user seeded: admin@fakulti.com")
        except Exception as e:
            logger.error(f"Failed to seed admin: {e}")
    
    product_count = await db.products.count_documents({})
    if product_count == 0:
        products = [
            {"id": str(uuid.uuid4()), "name": "Bombro - Bone Broth Hidrolizado", "code": "BOMBRO", "description": "Bone Broth Hidrolizado premium. Producto único en Ecuador. Rico en colágeno y nutrientes esenciales.", "price": 55.95, "original_price": 59.99, "image_url": "https://fakultisupplements.com/wp-content/uploads/2023/02/EDIT_BONE-BROTH-POWDER-200G_FAKULTI_2024.png", "stock": 150, "category": "nutrición", "active": True, "bot_config": {"personality": "Experto en nutrición y caldo de hueso hidrolizado. Apasionado por ayudar a las personas a mejorar su salud de forma natural.", "key_benefits": "Colágeno de alta absorción, mejora digestión, soporte articular, fuente de proteína, fácil de preparar como sopa caliente o fría", "usage_info": "Un sachet al día. Se puede tomar en el desayuno como sopa caliente o en la noche. Diluir en agua caliente y mezclar.", "restrictions": "No prometer curas. No decir que cura enfermedades. No afirmar que reemplaza tratamientos médicos. Nunca usar lenguaje médico complejo.", "faqs": "Se toma un sachet al día. Sabe como una sopa de hueso suave. Es apto para toda la familia. No contiene gluten."}},
            {"id": str(uuid.uuid4()), "name": "Gomitas Melatonina", "code": "GUMMELAT", "description": "Gomitas de melatonina para un descanso natural y reparador.", "price": 13.25, "original_price": 15.99, "image_url": "https://fakultisupplements.com/wp-content/uploads/2022/10/PRODUCTOS-FAKULTI-GUMMIES-DEFENSE.png", "stock": 200, "category": "bienestar", "active": True, "bot_config": {"personality": "Asesor de bienestar y descanso. Empático con las personas que tienen problemas de sueño.", "key_benefits": "Ayuda a conciliar el sueño de forma natural, sabor agradable, fácil de tomar, sin dependencia", "usage_info": "Tomar 1-2 gomitas 30 minutos antes de dormir. No exceder la dosis recomendada.", "restrictions": "No prometer que cura insomnio. No reemplaza tratamiento médico. No combinar con otros sedantes sin consultar médico.", "faqs": "Son gomitas con sabor frutal. Se toman antes de dormir. No generan dependencia. Aptas para adultos."}},
            {"id": str(uuid.uuid4()), "name": "CBD Colágeno Hidrolizado", "code": "CBD-COL", "description": "Colágeno hidrolizado con CBD para soporte articular y bienestar integral.", "price": 52.36, "original_price": 57.45, "image_url": "https://fakultisupplements.com/wp-content/uploads/2025/08/sachets-17.png", "stock": 100, "category": "cbd", "active": True, "bot_config": {"personality": "Especialista en bienestar integral y productos con CBD. Informado y objetivo sobre los beneficios del CBD.", "key_benefits": "Soporte articular, bienestar integral, colágeno para piel y articulaciones, CBD para relajación natural", "usage_info": "Un sachet al día diluido en agua. Preferiblemente en la mañana o antes de dormir.", "restrictions": "No prometer curas. CBD no es medicina. No reemplaza tratamientos médicos. Informar que el CBD es legal en Ecuador como suplemento.", "faqs": "El CBD es legal en Ecuador como suplemento. No produce efectos psicoactivos. Contiene colágeno hidrolizado más CBD."}},
            {"id": str(uuid.uuid4()), "name": "Pitch Up", "code": "PITCHUP", "description": "Suplemento energético natural para rendimiento físico y mental.", "price": 21.84, "original_price": 24.99, "image_url": "https://fakultisupplements.com/wp-content/uploads/2022/10/PRODUCTOS-FAKULTI-EXIT-FAT.png", "stock": 120, "category": "energía", "active": True, "bot_config": {"personality": "Coach de energía y rendimiento. Motivador y práctico.", "key_benefits": "Energía natural sin crash, mejora rendimiento físico y mental, ideal para deportistas y profesionales activos", "usage_info": "Tomar un sachet al día, preferiblemente en la mañana o antes de actividad física. Mezclar con agua.", "restrictions": "No prometer resultados deportivos específicos. No es un estimulante. No reemplaza una dieta balanceada.", "faqs": "No contiene cafeína artificial. Es energía natural. Se puede tomar todos los días."}},
            {"id": str(uuid.uuid4()), "name": "Magnesio Citrato", "code": "MAGCIT", "description": "Magnesio citrato de alta absorción para soporte muscular, nervioso y cardiovascular.", "price": 18.50, "original_price": 22.99, "image_url": "https://fakultisupplements.com/wp-content/uploads/2022/10/PRODUCTOS-FAKULTI-GUMMIES-DEFENSE.png", "stock": 180, "category": "bienestar", "active": True, "bot_config": {"personality": "Asesor de salud preventiva. Conocedor de los beneficios del magnesio para el bienestar diario.", "key_benefits": "Soporte muscular y nervioso, ayuda a relajación, contribuye al funcionamiento cardiovascular, alta absorción", "usage_info": "Un sachet al día diluido en agua. Se puede tomar en cualquier momento del día.", "restrictions": "No prometer curas. No reemplaza tratamientos médicos. Consultar con médico si toma otros medicamentos.", "faqs": "El magnesio citrato tiene mejor absorción que otras formas. Ayuda con calambres musculares. Seguro para uso diario."}},
        ]
        for p in products:
            p["created_at"] = datetime.now(timezone.utc).isoformat()
        await db.products.insert_many(products)
        logger.info("Products seeded")
    
    game_count = await db.games_config.count_documents({})
    if game_count == 0:
        games = [
            {
                "id": str(uuid.uuid4()),
                "game_type": "roulette",
                "name": "Ruleta Fakulti",
                "prizes": [
                    {"name": "10% Descuento", "probability": 30, "color": "#A3E635", "coupon": "RULETA10", "message": "Ganaste un 10% de descuento en tu proxima compra"},
                    {"name": "Envio Gratis", "probability": 20, "color": "#3B82F6", "coupon": "ENVIOGRATIS", "message": "Ganaste envio gratis en tu proxima compra"},
                    {"name": "Muestra Gratis Bombro", "probability": 10, "color": "#F59E0B", "coupon": "MUESTRA", "message": "Ganaste una muestra gratis de Bombro"},
                    {"name": "15% Descuento", "probability": 8, "color": "#8B5CF6", "coupon": "RULETA15", "message": "Ganaste un 15% de descuento"},
                    {"name": "2x1 Gomitas", "probability": 5, "color": "#EF4444", "coupon": "2X1GOMAS", "message": "Ganaste 2x1 en Gomitas Melatonina"},
                    {"name": "Sigue intentando", "probability": 27, "color": "#64748B", "coupon": "", "message": "No ganaste esta vez, pero tienes un cupon de 5%: FACULTY5"}
                ],
                "active": True,
                "max_plays_per_whatsapp": 1,
                "created_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": str(uuid.uuid4()),
                "game_type": "slot_machine",
                "name": "Tragamonedas Fakulti",
                "prizes": [
                    {"name": "20% Descuento", "probability": 10, "color": "#A3E635", "coupon": "SLOT20", "message": "Tres iguales! Ganaste un 20% de descuento"},
                    {"name": "15% Descuento", "probability": 15, "color": "#8B5CF6", "coupon": "SLOT15", "message": "Gran combinacion! 15% de descuento para ti"},
                    {"name": "10% Descuento", "probability": 25, "color": "#3B82F6", "coupon": "SLOT10", "message": "Buena jugada! 10% de descuento"},
                    {"name": "Envio Gratis", "probability": 20, "color": "#F59E0B", "coupon": "SLOTENVIO", "message": "Ganaste envio gratis en tu proxima compra"},
                    {"name": "5% Descuento", "probability": 30, "color": "#64748B", "coupon": "SLOT5", "message": "Tienes un 5% de descuento: SLOT5"}
                ],
                "active": True,
                "max_plays_per_whatsapp": 1,
                "created_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": str(uuid.uuid4()),
                "game_type": "scratch_card",
                "name": "Raspadita Fakulti",
                "prizes": [
                    {"name": "Descuento 25%", "probability": 5, "color": "#A3E635", "coupon": "RASPA25", "message": "Descubriste un 25% de descuento!"},
                    {"name": "Descuento 15%", "probability": 15, "color": "#8B5CF6", "coupon": "RASPA15", "message": "Debajo de la capa dorada: 15% de descuento"},
                    {"name": "Descuento 10%", "probability": 25, "color": "#3B82F6", "coupon": "RASPA10", "message": "Raspa y gana! 10% de descuento"},
                    {"name": "Envio Gratis", "probability": 25, "color": "#F59E0B", "coupon": "RASPAENVIO", "message": "Encontraste envio gratis"},
                    {"name": "5% Descuento", "probability": 30, "color": "#64748B", "coupon": "RASPA5", "message": "Un 5% de descuento te espera: RASPA5"}
                ],
                "active": True,
                "max_plays_per_whatsapp": 1,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
        ]
        await db.games_config.insert_many(games)
        logger.info("Game configs seeded")
    
    # Migrate old games to new ones
    old_mystery = await db.games_config.find_one({"game_type": "mystery_box"})
    if old_mystery:
        await db.games_config.delete_one({"game_type": "mystery_box"})
        logger.info("Removed mystery_box game config")
    old_lucky = await db.games_config.find_one({"game_type": "lucky_button"})
    if old_lucky:
        await db.games_config.delete_one({"game_type": "lucky_button"})
        logger.info("Removed lucky_button game config")
    if not await db.games_config.find_one({"game_type": "slot_machine"}):
        await db.games_config.insert_one({
            "id": str(uuid.uuid4()),
            "game_type": "slot_machine",
            "name": "Tragamonedas Fakulti",
            "prizes": [
                {"name": "20% Descuento", "probability": 10, "color": "#A3E635", "coupon": "SLOT20", "message": "Tres iguales! Ganaste un 20% de descuento"},
                {"name": "15% Descuento", "probability": 15, "color": "#8B5CF6", "coupon": "SLOT15", "message": "Gran combinacion! 15% de descuento para ti"},
                {"name": "10% Descuento", "probability": 25, "color": "#3B82F6", "coupon": "SLOT10", "message": "Buena jugada! 10% de descuento"},
                {"name": "Envio Gratis", "probability": 20, "color": "#F59E0B", "coupon": "SLOTENVIO", "message": "Ganaste envio gratis en tu proxima compra"},
                {"name": "5% Descuento", "probability": 30, "color": "#64748B", "coupon": "SLOT5", "message": "Tienes un 5% de descuento: SLOT5"}
            ],
            "active": True, "max_plays_per_whatsapp": 1, "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info("Slot machine game config seeded")
    if not await db.games_config.find_one({"game_type": "scratch_card"}):
        await db.games_config.insert_one({
            "id": str(uuid.uuid4()),
            "game_type": "scratch_card",
            "name": "Raspadita Fakulti",
            "prizes": [
                {"name": "Descuento 25%", "probability": 5, "color": "#A3E635", "coupon": "RASPA25", "message": "Descubriste un 25% de descuento!"},
                {"name": "Descuento 15%", "probability": 15, "color": "#8B5CF6", "coupon": "RASPA15", "message": "Debajo de la capa dorada: 15% de descuento"},
                {"name": "Descuento 10%", "probability": 25, "color": "#3B82F6", "coupon": "RASPA10", "message": "Raspa y gana! 10% de descuento"},
                {"name": "Envio Gratis", "probability": 25, "color": "#F59E0B", "coupon": "RASPAENVIO", "message": "Encontraste envio gratis"},
                {"name": "5% Descuento", "probability": 30, "color": "#64748B", "coupon": "RASPA5", "message": "Un 5% de descuento te espera: RASPA5"}
            ],
            "active": True, "max_plays_per_whatsapp": 1, "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info("Scratch card game config seeded")
    
    # Seed sample leads
    leads_count = await db.leads.count_documents({})
    if leads_count == 0:
        sample_leads = [
            {"id": str(uuid.uuid4()), "name": "Maria Garcia", "whatsapp": "0991234567", "city": "Quito", "email": "maria@email.com", "product_interest": "Bombro", "source": "TV", "season": "", "channel": "TV", "funnel_stage": "cliente_activo", "status": "activo", "game_used": "roulette", "prize_obtained": "10% Descuento", "purchase_history": [{"id": str(uuid.uuid4()), "product_name": "Bombro - Bone Broth Hidrolizado", "quantity": 2, "price": 111.90, "date": "2025-12-15"}], "coupon_used": "RULETA10", "recompra_date": "2026-01-15", "notes": "Cliente frecuente"},
            {"id": str(uuid.uuid4()), "name": "Carlos Lopez", "whatsapp": "0987654321", "city": "Guayaquil", "email": "carlos@email.com", "product_interest": "Colageno CBD", "source": "QR", "season": "", "channel": "QR", "funnel_stage": "interesado", "status": "activo", "game_used": "mystery_box", "prize_obtained": "Envio Gratis", "purchase_history": [], "coupon_used": None, "recompra_date": None, "notes": "Interesado en CBD"},
            {"id": str(uuid.uuid4()), "name": "Ana Martinez", "whatsapp": "0976543210", "city": "Cuenca", "email": "ana@email.com", "product_interest": "Gomitas Melatonina", "source": "Fibeca", "season": "", "channel": "Fibeca", "funnel_stage": "cliente_nuevo", "status": "activo", "game_used": None, "prize_obtained": None, "purchase_history": [{"id": str(uuid.uuid4()), "product_name": "Gomitas Melatonina", "quantity": 1, "price": 13.25, "date": "2026-01-20"}], "coupon_used": None, "recompra_date": "2026-02-20", "notes": "Compro en Fibeca"},
            {"id": str(uuid.uuid4()), "name": "Pedro Sanchez", "whatsapp": "0965432109", "city": "Quito", "email": "", "product_interest": "Pitch Up", "source": "pauta_digital", "season": "", "channel": "pauta_digital", "funnel_stage": "en_negociacion", "status": "activo", "game_used": None, "prize_obtained": None, "purchase_history": [], "coupon_used": None, "recompra_date": None, "notes": "Pidio cotizacion"},
            {"id": str(uuid.uuid4()), "name": "Laura Fernandez", "whatsapp": "0954321098", "city": "Guayaquil", "email": "laura@email.com", "product_interest": "Bombro", "source": "TV", "season": "", "channel": "TV", "funnel_stage": "nuevo", "status": "activo", "game_used": None, "prize_obtained": None, "purchase_history": [], "coupon_used": None, "recompra_date": None, "notes": ""},
            {"id": str(uuid.uuid4()), "name": "Roberto Diaz", "whatsapp": "0943210987", "city": "Ambato", "email": "", "product_interest": "", "source": "web", "season": "", "channel": "web", "funnel_stage": "perdido", "status": "inactivo", "game_used": "slot_machine", "prize_obtained": "5% Descuento", "purchase_history": [], "coupon_used": None, "recompra_date": None, "notes": "Sin respuesta despues de 3 recordatorios"},
        ]
        for lead in sample_leads:
            lead["last_interaction"] = datetime.now(timezone.utc).isoformat()
            lead["created_at"] = datetime.now(timezone.utc).isoformat()
        await db.leads.insert_many(sample_leads)
        logger.info("Sample leads seeded")

    await db.leads.create_index("whatsapp")
    await db.leads.create_index("funnel_stage")
    await db.leads.create_index("source")
    await db.game_plays.create_index("whatsapp")
    
    # Seed automation rules
    rules_count = await db.automation_rules.count_documents({})
    if rules_count == 0:
        default_rules = [
            {"id": str(uuid.uuid4()), "name": "Bienvenida automática", "trigger_type": "nuevo_lead", "trigger_value": "", "action_type": "respuesta_ia", "action_value": "Saluda al cliente con: Hola! Gracias por contactarnos. Soy el asesor virtual de Fakulti Laboratorios. Somos especialistas en suplementos naturales de alta calidad. Para brindarte una mejor atención, me podrías decir tu nombre?", "description": "Saluda automáticamente a cada nuevo lead que ingresa al sistema y solicita sus datos.", "active": True, "order": 1},
            {"id": str(uuid.uuid4()), "name": "Solicitar datos del cliente", "trigger_type": "lead_sin_datos", "trigger_value": "nombre,whatsapp,ciudad", "action_type": "respuesta_ia", "action_value": "Pregunta de forma natural los datos faltantes: nombre completo, número de WhatsApp, ciudad y producto de interés.", "description": "Cuando un lead no tiene datos completos, el bot los solicita durante la conversación.", "active": True, "order": 2},
            {"id": str(uuid.uuid4()), "name": "Clasificar etapa automáticamente", "trigger_type": "analisis_conversacion", "trigger_value": "", "action_type": "cambiar_etapa", "action_value": "Analiza keywords: precio/cuanto=interesado, cotiz/pago/envio=en_negociacion, comprar/confirmo=en_negociacion, no me interesa=perdido", "description": "Clasifica automáticamente la etapa del lead basándose en las palabras clave de la conversación.", "active": True, "order": 3},
            {"id": str(uuid.uuid4()), "name": "Primer recordatorio (4 horas)", "trigger_type": "sin_respuesta", "trigger_value": "4", "action_type": "enviar_mensaje", "action_value": "Hola, solo para saber si pudiste revisar la información que te envié. Si quieres te ayudo con más detalles sobre nuestros productos.", "description": "Envía un recordatorio amable después de 4 horas sin respuesta del lead.", "active": True, "order": 4},
            {"id": str(uuid.uuid4()), "name": "Segundo recordatorio (24 horas)", "trigger_type": "sin_respuesta", "trigger_value": "24", "action_type": "enviar_mensaje", "action_value": "Hola de nuevo, quería saber si aún tienes interés en los productos de Fakulti. Estoy aquí para ayudarte cuando lo necesites.", "description": "Segundo recordatorio después de 24 horas sin respuesta.", "active": True, "order": 5},
            {"id": str(uuid.uuid4()), "name": "Marcar como perdido", "trigger_type": "sin_respuesta", "trigger_value": "48", "action_type": "cambiar_etapa", "action_value": "perdido", "description": "Marca automáticamente al lead como perdido después de 48 horas sin respuesta.", "active": True, "order": 6},
            {"id": str(uuid.uuid4()), "name": "Transferir a humano", "trigger_type": "intencion_ia", "trigger_value": "queja,problema,reclamo,hablar con persona,agente", "action_type": "asignar_agente", "action_value": "Transfiere la conversación a un asesor humano cuando el bot detecta una queja, problema complejo o solicitud explícita de hablar con una persona.", "description": "Detecta intenciones críticas y deriva a un agente humano.", "active": True, "order": 7},
            {"id": str(uuid.uuid4()), "name": "Recomendación de producto", "trigger_type": "intencion_ia", "trigger_value": "consulta_producto,interes,salud,dolor,suplemento", "action_type": "respuesta_ia", "action_value": "Recomienda productos del catálogo de Fakulti basándose en las necesidades del cliente. Menciona beneficios sin hacer claims médicos.", "description": "Sugiere productos relevantes basados en el mensaje del cliente.", "active": True, "order": 8},
            {"id": str(uuid.uuid4()), "name": "Seguimiento post-compra", "trigger_type": "compra_realizada", "trigger_value": "", "action_type": "iniciar_secuencia", "action_value": "Inscribe automáticamente al cliente en la secuencia de fidelización del producto comprado.", "description": "Al registrar una compra, inicia la secuencia de mensajes de fidelización.", "active": True, "order": 9},
            {"id": str(uuid.uuid4()), "name": "Recordatorio de recompra (30 días)", "trigger_type": "dias_post_compra", "trigger_value": "30", "action_type": "enviar_mensaje", "action_value": "Hola! Ya pasó un mes desde tu última compra en Fakulti. Te gustaría repetir tu pedido? Tenemos nuevas promociones disponibles.", "description": "Envía un recordatorio para recompra 30 días después de la última compra.", "active": True, "order": 10},
        ]
        for r in default_rules:
            r["created_at"] = datetime.now(timezone.utc).isoformat()
        await db.automation_rules.insert_many(default_rules)
        logger.info("Automation rules seeded (10 rules)")
    
    # Seed default configs
    if not await db.whatsapp_config.find_one({"id": "main"}):
        await db.whatsapp_config.insert_one({"id": "main", "phone_number_id": "", "access_token": "", "verify_token": "fakulti-whatsapp-verify-token", "business_name": "Fakulti Laboratorios"})
    if not await db.ai_config.find_one({"id": "main"}):
        await db.ai_config.insert_one({"id": "main", "intent_analysis": True, "lead_classification": True, "product_recommendation": True, "suggested_responses": True})
    
    # Migrate old funnel stages
    migrated_caliente = await db.leads.update_many({"funnel_stage": "caliente"}, {"$set": {"funnel_stage": "en_negociacion"}})
    migrated_frio = await db.leads.update_many({"funnel_stage": "frio"}, {"$set": {"funnel_stage": "perdido"}})
    if migrated_caliente.modified_count or migrated_frio.modified_count:
        logger.info(f"Migrated stages: caliente->en_negociacion ({migrated_caliente.modified_count}), frio->perdido ({migrated_frio.modified_count})")
    
    # ===== COMPREHENSIVE PHONE NORMALIZATION & LEAD DEDUP =====
    all_leads = await db.leads.find({}, {"_id": 0}).to_list(10000)
    phone_groups = {}
    for lead in all_leads:
        raw_phone = lead.get("whatsapp", "")
        if not raw_phone:
            continue
        norm = normalize_phone_ec(raw_phone)
        if norm not in phone_groups:
            phone_groups[norm] = []
        phone_groups[norm].append(lead)
    
    merged_count = 0
    normalized_count = 0
    for norm_phone, leads_group in phone_groups.items():
        if len(leads_group) == 1:
            lead = leads_group[0]
            if lead.get("whatsapp") != norm_phone:
                old_phone = lead["whatsapp"]
                await db.leads.update_one({"id": lead["id"]}, {"$set": {"whatsapp": norm_phone}})
                # Update session meta and messages that reference this lead by old phone session
                old_session = f"wa_{old_phone}"
                new_session = f"wa_{norm_phone}"
                if old_session != new_session:
                    await db.chat_sessions_meta.update_many({"session_id": old_session}, {"$set": {"session_id": new_session}})
                    await db.chat_messages.update_many({"session_id": old_session}, {"$set": {"session_id": new_session}})
                normalized_count += 1
        else:
            # Multiple leads with the same normalized phone - merge them
            # Pick the primary lead (the one with most data: has name, has purchase_history, most recent interaction)
            leads_group.sort(key=lambda lead_item: (
                bool(lead_item.get("name", "").strip()),
                len(lead_item.get("purchase_history", [])),
                lead_item.get("last_interaction", ""),
            ), reverse=True)
            primary = leads_group[0]
            primary_id = primary["id"]
            
            for dup in leads_group[1:]:
                dup_id = dup["id"]
                # Merge purchase_history
                dup_purchases = dup.get("purchase_history", [])
                if dup_purchases:
                    await db.leads.update_one({"id": primary_id}, {"$push": {"purchase_history": {"$each": dup_purchases}}})
                # Fill missing fields from duplicate
                update_fields = {}
                for field in ["name", "city", "email", "product_interest", "source", "channel", "game_used", "prize_obtained", "coupon_used"]:
                    if not primary.get(field) and dup.get(field):
                        update_fields[field] = dup[field]
                if dup.get("assigned_advisor") and not primary.get("assigned_advisor"):
                    update_fields["assigned_advisor"] = dup["assigned_advisor"]
                if update_fields:
                    await db.leads.update_one({"id": primary_id}, {"$set": update_fields})
                
                # Transfer chat messages and sessions from duplicate to primary
                old_sessions_for_dup = [f"wa_{dup.get('whatsapp', '')}", f"wa_{normalize_phone_ec(dup.get('whatsapp', ''))}"]
                for old_sid in old_sessions_for_dup:
                    await db.chat_messages.update_many({"lead_id": dup_id}, {"$set": {"lead_id": primary_id}})
                    await db.chat_sessions_meta.update_many({"lead_id": dup_id}, {"$set": {"lead_id": primary_id, "lead_name": primary.get("name", "")}})
                
                # Delete the duplicate lead
                await db.leads.delete_one({"id": dup_id})
                logger.info(f"Merged duplicate lead {dup.get('name','')} ({dup.get('whatsapp','')}) into {primary.get('name','')} ({primary.get('whatsapp','')})")
            
            # Normalize primary phone and update session IDs
            old_phone = primary.get("whatsapp", "")
            if old_phone != norm_phone:
                await db.leads.update_one({"id": primary_id}, {"$set": {"whatsapp": norm_phone}})
            
            # Consolidate all session variants to normalized session ID
            all_variants = set()
            for dup_lead in leads_group:
                p = dup_lead.get("whatsapp", "")
                all_variants.add(f"wa_{p}")
                all_variants.add(f"wa_{normalize_phone_ec(p)}")
                all_variants.add(f"wa_{phone_to_international(p)}")
            target_session = f"wa_{norm_phone}"
            for old_sid in all_variants:
                if old_sid != target_session:
                    await db.chat_sessions_meta.update_many({"session_id": old_sid}, {"$set": {"session_id": target_session, "lead_name": primary.get("name", "")}})
                    await db.chat_messages.update_many({"session_id": old_sid}, {"$set": {"session_id": target_session, "lead_id": primary_id}})
            
            merged_count += len(leads_group) - 1
    
    if normalized_count or merged_count:
        logger.info(f"Phone migration: {normalized_count} normalized, {merged_count} duplicates merged")
    
    # Ensure season and channel fields exist on all leads
    await db.leads.update_many({"season": {"$exists": False}}, {"$set": {"season": ""}})
    await db.leads.update_many({"channel": {"$exists": False}}, {"$set": {"channel": ""}})
    
    # Deactivate games except slot machine (standby mode)
    await db.games_config.update_many({"game_type": {"$in": ["roulette", "scratch_card"]}}, {"$set": {"active": False}})
    logger.info("Games standby: only slot_machine active")
    
    # Seed default QR campaigns
    qr_count = await db.qr_campaigns.count_documents({})
    if qr_count == 0:
        default_qr_campaigns = [
            {
                "id": str(uuid.uuid4()),
                "name": "TV - Anuncio General",
                "channel": "TV/QR",
                "source": "TV",
                "product": "",
                "initial_message": "Hola, vi esto en TV",
                "description": "QR para anuncios de televisión. El cliente escanea el QR que aparece en pantalla.",
                "active": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": str(uuid.uuid4()),
                "name": "Fibeca - Punto de Venta",
                "channel": "Fibeca",
                "source": "Fibeca",
                "product": "",
                "initial_message": "Hola, los vi en Fibeca",
                "description": "QR para puntos de venta en Fibeca.",
                "active": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": str(uuid.uuid4()),
                "name": "Evento - Feria de Salud",
                "channel": "Evento",
                "source": "Evento",
                "product": "Bombro",
                "initial_message": "Hola, los conoci en la feria",
                "description": "QR para ferias y eventos presenciales.",
                "active": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            },
        ]
        await db.qr_campaigns.insert_many(default_qr_campaigns)
        logger.info("Default QR campaigns seeded")
    
    # Ensure scan_count on QR campaigns
    await db.qr_campaigns.update_many({"scan_count": {"$exists": False}}, {"$set": {"scan_count": 0}})
    
    # Ensure bot_paused field on leads
    await db.leads.update_many({"bot_paused": {"$exists": False}}, {"$set": {"bot_paused": False}})
    
    # Ensure assigned_advisor field on leads
    await db.leads.update_many({"assigned_advisor": {"$exists": False}}, {"$set": {"assigned_advisor": ""}})
    
    # Ensure bot_config on products
    products_without_bot = await db.products.find({"bot_config": {"$exists": False}}, {"_id": 0, "id": 1, "name": 1, "description": 1}).to_list(100)
    for p in products_without_bot:
        default_bot = {
            "personality": f"Especialista en {p.get('name', 'el producto')}. Amigable y profesional.",
            "key_benefits": p.get("description", ""),
            "usage_info": "Consultar indicaciones del producto.",
            "restrictions": "No hacer promesas medicas. No afirmar que cura enfermedades.",
            "faqs": ""
        }
        await db.products.update_one({"id": p["id"]}, {"$set": {"bot_config": default_bot}})
    if products_without_bot:
        logger.info(f"Added bot_config to {len(products_without_bot)} products")
    
    logger.info("Fakulti CRM Backend ready")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
